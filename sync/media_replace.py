"""
Targeted media replacement — redo the audio and/or image for specific words, driven by a
backlog sheet (data/media_replacements.xlsx) you add rows to whenever you notice a bad clip
or picture. The tool records INTENT and previews; the existing pipelines do the heavy work.

Sheet columns (Word + Type are yours; blue-headed columns are tool-managed):
  Word, Type                 identify the word (article prefix ok: "die Stadt"); Type is
                             noun/verb/adjective/adverb and picks the vocabulary table
  Level (auto)               filled by the tool on resolution; fill it yourself ONLY to pin
                             one of several levels when the tool reports an ambiguity
  Replace_Audio, Replace_Image   mark either or both with "x"
  Audio_Variants             which clips: all (default) / singular / plural / sentence / a+b
  Voice                      optional exact Azure voice pin (default: each take ROTATES to a
                             different voice — same text + same neural voice would reproduce
                             the same bad clip)
  Pronunciation_Hint         optional respelling substituted for the word in the spoken text
  Image_Note                 optional feedback appended to the image-generation prompt
  Status                     tool-managed lifecycle; CLEAR it to request another round

Lifecycle per row (Status text mirrors it):
  audio:  preview (listen in data/media_preview) → --approve → approved (committed to
          audio_overrides.json) → run audio_sync.py → published ✓
  image:  queued (decision released, zero-gap: the current image keeps shipping) → run
          image_sync.py → awaiting review → image_review.py pick → done ✓

The tool never touches R2 and never publishes: audio ships via audio_sync.py, images via the
normal image pipeline. Approved audio takes live in sync/audio_overrides.json — COMMIT that
file, or another machine will revert the replacements on its next audio_sync run.

Usage:
  python media_replace.py                 # process the sheet: resolve, preview, track
  python media_replace.py --approve       # approve ALL previewed audio takes
  python media_replace.py --approve Hund "die Stadt"   # approve specific words only
  python media_replace.py --dry-run       # report what would happen; change nothing
  python media_replace.py --audio-only | --images-only
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

import audio_engine
import audio_overrides
import audio_sync
import image_decisions
import sync

logger = logging.getLogger("media_replace")

SHEET_PATH = sync.DATA_DIR / "media_replacements.xlsx"
PREVIEW_DIR = sync.DATA_DIR / "media_preview"
STATE_PATH = PREVIEW_DIR / "state.json"   # uncommitted sidecar: preview takes + image request markers

TYPE_TO_TABLE = {"noun": "nouns", "verb": "verbs", "adjective": "adverbs_adjectives", "adverb": "adverbs_adjectives"}
# Tables whose words can carry images today. Deliberately config, not an assumption baked into
# logic: when the image pipeline learns other word types, extend this set and nothing else here.
IMAGE_CAPABLE_TABLES = {"nouns"}
VARIANTS = ("singular", "plural", "sentence")

# Canonical column keys -> acceptable header spellings (matched case-insensitively, the
# " (auto)" suffix ignored) so a cosmetic header edit doesn't break the tool.
HEADERS = {
    "word": "Word", "type": "Type", "level": "Level", "replace_audio": "Replace_Audio",
    "replace_image": "Replace_Image", "audio_variants": "Audio_Variants", "voice": "Voice",
    "hint": "Pronunciation_Hint", "image_note": "Image_Note", "status": "Status",
}
_ARTICLES = ("der", "die", "das")


# ---------------------------------------------------------------------------
# Sidecar state (local; recomputable except preview take counters)
# ---------------------------------------------------------------------------

def _load_state() -> dict[str, Any]:
    try:
        return json.loads(STATE_PATH.read_text("utf-8"))
    except FileNotFoundError:
        return {}
    except Exception:  # noqa: BLE001 — worst case: a preview regenerates / a done-marker is re-derived
        logger.warning("preview state unreadable — starting fresh (previews may regenerate)")
        return {}


def _save_state(state: dict[str, Any]) -> None:
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)
    STATE_PATH.write_text(json.dumps(state, indent=2, sort_keys=True, ensure_ascii=False) + "\n", "utf-8")


# ---------------------------------------------------------------------------
# Vocabulary resolution
# ---------------------------------------------------------------------------

def _norm(text: str | None) -> str:
    return (text or "").replace("\xa0", " ").strip()


def _word_key(word: str, type_: str) -> str:
    """Lookup key for a target word: lowercased, article-stripped for nouns."""
    w = _norm(word).lower()
    if type_ == "noun":
        head, sep, tail = w.partition(" ")
        if sep and head in _ARTICLES and tail:
            w = tail
    return w


def load_vocabulary() -> tuple[dict[str, list[dict[str, Any]]], dict[tuple[str, str], list[dict[str, Any]]]]:
    """Read all tables (tolerating invalid rows — they just can't be replacement targets)
    and index rows by (word_key, type). Aborts on cross-table id collisions, which would
    make replacement state ambiguous everywhere ids are global."""
    rows_by_table: dict[str, list[dict[str, Any]]] = {}
    index: dict[tuple[str, str], list[dict[str, Any]]] = {}
    # Vocabulary hygiene is sync.py's job — don't repeat its per-row warning wall here.
    # A replacement row targeting an invalid word still fails with its own clear error.
    sync_logger = logging.getLogger("sync")
    prev_level = sync_logger.level
    sync_logger.setLevel(logging.ERROR)
    try:
        for table in sync.TABLE_CONFIG:
            rows, _, _ = sync.read_excel(table, skip_invalid=True)
            rows_by_table[table] = rows
    finally:
        sync_logger.setLevel(prev_level)
    for table, rows in rows_by_table.items():
        for row in rows:
            key = ((row.get("word") or "").lower(), (row.get("type") or "").lower())
            index.setdefault(key, []).append(row)
    problems = sync.find_cross_table_id_collisions(rows_by_table)
    if problems:
        raise sync.ValidationError(
            "cross-table id collision(s) — fix the vocabulary first:\n  " + "\n  ".join(problems)
        )
    return rows_by_table, index


def resolve(plan: dict[str, Any], index: dict[tuple[str, str], list[dict[str, Any]]]) -> None:
    """Fill plan['vrow'] (the vocabulary row) or plan['errors']. A pre-filled Level pins the
    choice among same-word-same-type candidates at several levels."""
    candidates = index.get((_word_key(plan["word"], plan["type"]), plan["type"]), [])
    if plan["level"]:
        wanted = plan["level"].upper()
        candidates = [r for r in candidates if (r.get("level") or "").upper() == wanted]
    if not candidates:
        plan["errors"].append(
            f"word not found: {plan['word']!r} ({plan['type']}"
            + (f", level {plan['level']}" if plan["level"] else "")
            + ") — check spelling/type; rows invalid in the vocab sheet cannot be targeted"
        )
    elif len(candidates) > 1:
        levels = ", ".join(sorted(r["level"] for r in candidates))
        plan["errors"].append(f"ambiguous — exists at levels {levels}; fill Level to pin one")
    else:
        plan["vrow"] = candidates[0]


# ---------------------------------------------------------------------------
# Sheet reading
# ---------------------------------------------------------------------------

def _read_sheet(path: Path):
    """Open the backlog workbook read-write and map headers by name.
    Returns (workbook, worksheet, {canonical_key: column_index})."""
    if not path.exists():
        logger.error("Backlog sheet not found: %s", path)
        sys.exit(1)
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as exc:  # noqa: BLE001 — e.g. open in Excel with a lock
        logger.error("Cannot open %s (%s) — close it in Excel/Numbers and retry.", path.name, exc)
        sys.exit(1)
    ws = wb.active
    cols: dict[str, int] = {}
    for j, cell in enumerate(ws[1], start=1):
        name = _norm(str(cell.value) if cell.value is not None else "")
        name = re.sub(r"\s*\(auto\)\s*$", "", name, flags=re.IGNORECASE).lower()
        for key, canonical in HEADERS.items():
            if name == canonical.lower():
                cols[key] = j
    missing = [HEADERS[k] for k in HEADERS if k not in cols]
    if missing:
        logger.error("%s is missing column(s): %s", path.name, ", ".join(missing))
        sys.exit(1)
    return wb, ws, cols


def _build_plans(ws, cols: dict[str, int], index) -> list[dict[str, Any]]:
    """One plan per non-empty sheet row: parsed fields + resolution + validation errors."""
    plans: list[dict[str, Any]] = []
    claimed: dict[tuple[str, str], int] = {}   # (word_id, part) -> first sheet row that owns it
    for r in range(2, ws.max_row + 1):
        cell = lambda key: _norm(str(ws.cell(r, cols[key]).value)) if ws.cell(r, cols[key]).value is not None else ""  # noqa: E731
        word = cell("word")
        if not word:
            continue
        plan: dict[str, Any] = {
            "row": r, "word": word, "type": cell("type").lower(), "level": cell("level"),
            "do_audio": bool(cell("replace_audio")), "do_image": bool(cell("replace_image")),
            "variants_raw": cell("audio_variants").lower(), "voice": cell("voice"),
            "hint": cell("hint"), "note": cell("image_note"), "status": cell("status"),
            "errors": [], "vrow": None, "variants": list(VARIANTS),
        }
        plans.append(plan)

        if plan["type"] not in TYPE_TO_TABLE:
            plan["errors"].append(f"invalid Type {plan['type']!r} — use noun/verb/adjective/adverb")
            continue
        plan["table"] = TYPE_TO_TABLE[plan["type"]]
        if not plan["do_audio"] and not plan["do_image"]:
            plan["errors"].append("mark Replace_Audio and/or Replace_Image")
        if plan["do_image"] and plan["table"] not in IMAGE_CAPABLE_TABLES:
            plan["errors"].append(f"images are not enabled for {plan['type']}s yet (nouns only today)")
        if plan["variants_raw"] and plan["variants_raw"] != "all":
            parts = [v.strip() for v in plan["variants_raw"].split("+")]
            bad = [v for v in parts if v not in VARIANTS]
            if bad:
                plan["errors"].append(f"invalid Audio_Variants {plan['variants_raw']!r}")
            else:
                plan["variants"] = parts
        if plan["voice"]:
            if plan["voice"] not in audio_engine.ALL_VOICES:
                plan["errors"].append(f"unknown Voice {plan['voice']!r} — see the pools in audio_engine.py")
            elif plan["voice"] in audio_engine.DISABLED_VOICES:
                plan["errors"].append(f"Voice {plan['voice']!r} is disabled")
        if plan["errors"]:
            continue

        resolve(plan, index)
        if plan["vrow"] is None:
            continue
        vrow = plan["vrow"]
        plan["wid"] = vrow["id"]
        # Explicitly requested variants must exist on the word; 'all' just takes what's there.
        if plan["variants_raw"] and plan["variants_raw"] != "all" and plan["do_audio"]:
            if "plural" in plan["variants"] and (plan["table"] != "nouns" or not vrow.get("plural")):
                plan["errors"].append("no plural clip exists for this word")
            if "sentence" in plan["variants"] and not vrow.get("german_sentence"):
                plan["errors"].append("no sentence clip exists for this word (German sentence empty)")
        for part in ("audio", "image"):
            if plan[f"do_{part}"]:
                owner = claimed.setdefault((plan["wid"], part), r)
                if owner != r:
                    plan["errors"].append(f"duplicate: row {owner} already targets this word's {part}")
    return plans


# ---------------------------------------------------------------------------
# Audio: preview / approve / track
# ---------------------------------------------------------------------------

def _safe_name(word: str) -> str:
    return re.sub(r"[^\w\-]+", "_", word, flags=re.UNICODE).strip("_")


def _short_voice(voice: str) -> str:
    return voice.replace("de-DE-", "").replace("Neural", "")


def _target_descriptors(plan: dict[str, Any]) -> list[dict[str, Any]]:
    """The clips this row targets, built by the SAME code the real sync uses (pre-override)."""
    descs = audio_sync._row_descriptors(plan["table"], plan["vrow"])
    return [d for d in descs if d["variant"] in plan["variants"]]


def _overridden(plan: dict[str, Any], desc: dict[str, Any], rec: dict[str, Any]) -> dict[str, Any]:
    """A copy of `desc` with the override record applied — exactly what audio_sync will build."""
    d = dict(desc)
    pool = audio_engine.voice_pool_for(plan["table"], plan["vrow"], d["variant"])
    audio_overrides.apply(d, rec, pool, _norm(plan["vrow"].get("word")))
    return d


def _override_rec(plan: dict[str, Any], take: int) -> dict[str, Any]:
    rec: dict[str, Any] = {"take": take}
    if plan["voice"]:
        rec["voice"] = plan["voice"]
    if plan["hint"]:
        rec["hint"] = plan["hint"]
    return rec


def _next_take(plan: dict[str, Any], committed: audio_overrides.Store, srec: dict[str, Any]) -> int:
    """One past the highest take this word has seen — committed or merely previewed — so a
    rejected preview is never repeated and the voice rotation always advances."""
    takes = [int((committed.get(d["id"]) or {}).get("take") or 0) for d in _target_descriptors(plan)]
    takes.append(int((srec.get("audio") or {}).get("take") or 0))
    return max(takes) + 1


def _preview_audio(plan: dict[str, Any], committed: audio_overrides.Store, srec: dict[str, Any]) -> str:
    """Synthesize preview clips for the next take into PREVIEW_DIR. Returns the Status text."""
    take = _next_take(plan, committed, srec)
    rec = _override_rec(plan, take)
    safe = _safe_name(plan["word"])
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)
    for old in PREVIEW_DIR.glob(f"{safe}__*.m4a"):
        old.unlink(missing_ok=True)
    clips: dict[str, dict[str, Any]] = {}
    voices = set()
    for desc in _target_descriptors(plan):
        d = _overridden(plan, desc, rec)
        path = PREVIEW_DIR / f"{safe}__{d['variant']}_take{take}_{_short_voice(d['voice'])}.m4a"
        audio_engine.synthesize(d["text"], d["voice"], path)
        clips[d["id"]] = {"variant": d["variant"], "voice": d["voice"], "hash": d["audio_hash"],
                          "text": d["text"], "file": path.name}
        voices.add(_short_voice(d["voice"]))
    srec["audio"] = {"take": take, "clips": clips}
    return (f"audio: PREVIEW take {take} ({', '.join(sorted(voices))}) — "
            f"listen in data/media_preview, then media_replace.py --approve")


def _approve_audio(plan: dict[str, Any], committed: audio_overrides.Store, srec: dict[str, Any]) -> str:
    """Commit the previewed take to audio_overrides.json. Refuses if the vocabulary row
    changed since the preview (the published clip would not be the one heard)."""
    preview = srec.get("audio") or {}
    take = int(preview.get("take") or 0)
    rec = _override_rec(plan, take)
    for desc in _target_descriptors(plan):
        d = _overridden(plan, desc, rec)
        heard = (preview.get("clips") or {}).get(d["id"])
        if not heard or heard.get("hash") != d["audio_hash"]:
            return ("audio: ERROR — vocabulary changed since the preview was made; "
                    "clear Status to generate a fresh preview")
    for did in preview["clips"]:
        committed[did] = dict(rec)
    safe = _safe_name(plan["word"])
    for old in PREVIEW_DIR.glob(f"{safe}__*.m4a"):
        old.unlink(missing_ok=True)
    srec.pop("audio", None)
    srec["audio_done"] = {"take": take}
    return f"audio: approved take {take} — run audio_sync.py to publish"


def _track_audio(plan: dict[str, Any], committed: audio_overrides.Store,
                 srec: dict[str, Any], cache_index: dict[str, str]) -> str:
    """Status for a row whose audio request is already in flight."""
    if srec.get("audio"):
        take = srec["audio"]["take"]
        voices = ", ".join(sorted({_short_voice(c["voice"]) for c in srec["audio"]["clips"].values()}))
        missing = [c["file"] for c in srec["audio"]["clips"].values() if not (PREVIEW_DIR / c["file"]).exists()]
        if missing:
            return "audio: preview files lost — clear Status to regenerate"
        return (f"audio: PREVIEW take {take} ({voices}) — "
                f"listen in data/media_preview, then media_replace.py --approve")
    # Approved (or state lost after approval): the committed overrides are the truth —
    # rebuild the expected recipes from them and compare against the synthesis cache.
    expected: list[dict[str, Any]] = []
    take = 0
    for desc in _target_descriptors(plan):
        rec = committed.get(desc["id"])
        if rec is None:
            return "audio: no active replacement recorded — clear Status to restart"
        take = max(take, int(rec.get("take") or 0))
        expected.append(_overridden(plan, desc, rec))
    if expected and all(cache_index.get(d["id"]) == d["audio_hash"] for d in expected):
        return "audio: published ✓"
    return f"audio: approved take {take} — run audio_sync.py to publish"


# ---------------------------------------------------------------------------
# Image: request / track
# ---------------------------------------------------------------------------

def _request_image(plan: dict[str, Any], store, queue, opts, srec: dict[str, Any]) -> str:
    prev = image_decisions.request_replacement(store, plan["wid"])
    queue.pop(plan["wid"], None)   # any stale candidates from an old round: force a fresh one
    if plan["note"]:
        image_decisions.set_note(opts, plan["wid"], plan["note"])
    srec["image"] = {"requested": True, "prev_hash": prev}
    return "image: queued — run image_sync.py to generate candidates"


def _track_image(plan: dict[str, Any], store, queue, srec: dict[str, Any]) -> str:
    rec = store.get(plan["wid"])
    prev_hash = (srec.get("image") or {}).get("prev_hash")
    if rec is None:
        return "image: queued — run image_sync.py to generate candidates"
    if rec.get("replace_requested"):
        if plan["wid"] in queue:
            return "image: awaiting review — run image_review.py"
        return "image: queued — run image_sync.py to generate candidates"
    status = rec.get("status")
    if status == "review":
        return "image: awaiting review — run image_review.py"
    if status == "approved":
        if prev_hash is not None and rec.get("content_hash") == prev_hash:
            return "image: kept current (replacement round found nothing usable)"
        return "image: done ✓"
    if status == "none":
        return "image: no image (reviewer chose none)"
    return "image: queued — run image_sync.py to generate candidates"


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def _require_azure_speech() -> None:
    missing = []
    if not os.environ.get("AZURE_SPEECH_KEY"):
        missing.append("AZURE_SPEECH_KEY")
    if not os.environ.get("AZURE_SPEECH_ENDPOINT") and not os.environ.get("AZURE_SPEECH_REGION"):
        missing.append("AZURE_SPEECH_ENDPOINT|AZURE_SPEECH_REGION")
    if missing:
        logger.error("Audio previews need Azure Speech credentials in sync/.env: %s", ", ".join(missing))
        sys.exit(1)


def main() -> None:
    parser = argparse.ArgumentParser(description="Replace audio/images for specific words (backlog-driven).")
    parser.add_argument("--approve", nargs="*", metavar="WORD",
                        help="Approve previewed audio takes (all previews, or only the named words).")
    parser.add_argument("--dry-run", action="store_true", help="Report what would happen; change nothing.")
    scope = parser.add_mutually_exclusive_group()
    scope.add_argument("--audio-only", action="store_true", help="Process only Replace_Audio marks.")
    scope.add_argument("--images-only", action="store_true", help="Process only Replace_Image marks.")
    verbosity = parser.add_mutually_exclusive_group()
    verbosity.add_argument("-v", "--verbose", action="store_true")
    verbosity.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args()
    sync._setup_logging(args.verbose, args.quiet)

    logger.info("Reading vocabulary…")
    try:
        _, index = load_vocabulary()
    except sync.ValidationError as exc:
        logger.error("%s", exc)
        sys.exit(1)

    wb, ws, cols = _read_sheet(SHEET_PATH)
    plans = _build_plans(ws, cols, index)
    if not plans:
        logger.info("Backlog sheet has no rows — nothing to do.")
        return

    state = _load_state()
    committed = audio_overrides.load()
    store = image_decisions.load()
    queue = image_decisions.load_review_queue()
    opts = image_decisions.load_prompt_opts()
    cache_index: dict[str, str] = {}
    try:
        cache_index = json.loads(audio_sync.INDEX_PATH.read_text())
    except Exception:  # noqa: BLE001 — no cache yet: nothing counts as published
        pass

    approve_words = None
    if args.approve is not None:
        # Normalize like the resolver (case, articles) so "die Stadt" and "Stadt" both match.
        approve_words = {_word_key(w, "noun") for w in args.approve} or None  # empty = approve all

    # Plan first: any empty-Status audio row will synthesize a preview this run, so the Azure
    # credential check must pass before ANY state is touched.
    need_synth = [p for p in plans
                  if not p["errors"] and p["do_audio"] and not args.images_only and not p["status"]]
    if need_synth and not args.dry_run:
        _require_azure_speech()

    audio_dirty = image_dirty = False
    errors = approved = 0
    for plan in plans:
        if plan["errors"]:
            plan["status_out"] = "ERROR: " + "; ".join(plan["errors"])
            errors += 1
            continue
        srec = state.setdefault(plan["wid"], {"word": plan["word"], "type": plan["type"]})
        # The sheet contract: an EMPTY Status cell means "start (or restart) this request" —
        # a fresh row, a rejected preview, or a re-replacement of an already-replaced word.
        fresh = not plan["status"]
        parts: list[str] = []

        if plan["do_audio"] and not args.images_only:
            in_approve = (args.approve is not None and srec.get("audio")
                          and (approve_words is None
                               or _word_key(plan["word"], plan["type"]) in approve_words))
            if args.dry_run:
                if in_approve:
                    parts.append(_track_audio(plan, committed, srec, cache_index) + " [would approve]")
                elif fresh:
                    take = _next_take(plan, committed, srec)
                    rec = _override_rec(plan, take)
                    voices = sorted({_short_voice(_overridden(plan, d, rec)["voice"])
                                     for d in _target_descriptors(plan)})
                    parts.append(f"audio: would preview take {take} ({', '.join(voices)})")
                else:
                    parts.append(_track_audio(plan, committed, srec, cache_index))
            elif in_approve:
                text = _approve_audio(plan, committed, srec)
                audio_dirty = True
                parts.append(text)
                if text.startswith("audio: ERROR"):
                    errors += 1
                else:
                    approved += 1
            elif fresh:
                srec.pop("audio_done", None)   # restarting after a completed round
                parts.append(_preview_audio(plan, committed, srec))
            else:
                parts.append(_track_audio(plan, committed, srec, cache_index))

        if plan["do_image"] and not args.audio_only:
            if args.dry_run:
                parts.append("image: would queue for regeneration" if fresh
                             else _track_image(plan, store, queue, srec))
            elif fresh:
                parts.append(_request_image(plan, store, queue, opts, srec))
                image_dirty = True
            else:
                parts.append(_track_image(plan, store, queue, srec))

        plan["status_out"] = " | ".join(parts) if parts else plan["status"]

    if args.approve is not None and not args.dry_run and not approved:
        logger.warning("--approve matched no previewed rows (previews are made by a plain run first).")

    # Persist everything (skip in dry-run): stores first, then the sheet.
    if not args.dry_run:
        if audio_dirty:
            audio_overrides.save(committed)
        if image_dirty:
            image_decisions.save(store)
            image_decisions.save_review_queue(queue)
            image_decisions.save_prompt_opts(opts)
        _save_state(state)
        for plan in plans:
            if plan.get("vrow") is not None and not _norm(str(ws.cell(plan["row"], cols["level"]).value or "")):
                ws.cell(plan["row"], cols["level"]).value = plan["vrow"]["level"]
            if plan.get("status_out"):
                ws.cell(plan["row"], cols["status"]).value = plan["status_out"]
        try:
            wb.save(SHEET_PATH)
        except PermissionError:
            logger.error("Cannot write %s — close it in Excel/Numbers and re-run. "
                         "(State was saved; re-running is safe.)", SHEET_PATH.name)
            sys.exit(1)

    label = "[DRY RUN] " if args.dry_run else ""
    print(f"\n{'─' * 60}\n{label}Replacement backlog ({len(plans)} row(s))")
    for plan in plans:
        print(f"  {plan['word']:<24} {plan.get('status_out') or plan['status'] or '(no action)'}")
    print("─" * 60)
    if audio_dirty:
        print("Approved audio takes written to sync/audio_overrides.json — COMMIT this file.")
    if errors:
        sys.exit(1)


if __name__ == "__main__":
    main()
