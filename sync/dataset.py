"""V2 dataset reader (WD-ID, LG-FR-9..14): spreadsheet → core rows + translation
rows + id aliases, registry-driven.

One read of a table's sheet produces:
  core         — the German word rows (id = v2 identity, level-free)
  translations — one row per word × language that has content (id = "<word_id>:<lang>")
  aliases      — v1 id → v2 id rows for the re-key migration (id = legacy id)
  protected    — v2 ids of rows skipped by --skip-invalid (diff must not delete them)
  coverage     — per-language translated-row counts for the publish preview (LG-FR-14)

Variant fallback (LG-FR-12) is resolved AT PUBLISH TIME into the variant's rows:
a variant row is emitted only where the variant differs from its base; the worker
serves base-then-variant overlay. Here we emit variant rows exactly as authored
(sparse), and the coverage report counts a variant as covered when its base is.

All validation errors carry the sheet row number and the word, so the operator
can find the cell. Two rows collapsing to one v2 id is an ERROR naming both rows
and the Sense column as the fix (homonyms, WD-ID-3).
"""

from __future__ import annotations

import hashlib
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

import registry
from registry import (
    GERMAN_ARTICLES, LANGUAGES, REQUIRED_LANGUAGE, TABLES, TRANSLATION_FIELDS,
    VALID_SENSE, Table, compute_legacy_id, compute_word_id, is_ignored_header,
    known_headers_for,
)

logger = logging.getLogger("sync")

DATA_DIR = Path(__file__).parent.parent / "data"

# A1, A2, B1, B2, C1, C2 with optional .1 or .2 sub-level (same rule as v1)
VALID_LEVEL = re.compile(r"^(A1|A2|B1|B2|C1|C2)(\.[12])?$")

_NOUN_ARTICLES = GERMAN_ARTICLES


class ValidationError(Exception):
    """Structural or row-level validation failure."""


@dataclass
class TableDataset:
    core: list[dict[str, Any]] = field(default_factory=list)
    translations: list[dict[str, Any]] = field(default_factory=list)
    aliases: list[dict[str, Any]] = field(default_factory=list)
    protected: set[str] = field(default_factory=set)
    protected_aliases: set[str] = field(default_factory=set)  # legacy ids of skipped rows
    skipped: int = 0
    coverage: dict[str, int] = field(default_factory=dict)  # lang -> rows with content


# ---------------------------------------------------------------------------
# Cell helpers (v1 semantics preserved)
# ---------------------------------------------------------------------------

def _clean(value: Any) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str):
        value = str(value)
    return value.replace("\xa0", " ").strip() or None


_TRUTHY = {"1", "true", "yes", "x", "y"}


def _to_bool01(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return 1 if value != 0 else 0
    return 1 if str(value).strip().lower() in _TRUTHY else 0


def _capitalize_noun(value: str) -> str:
    """Capitalize a German noun's first letter, preserving an optional leading
    article ('die hunde' -> 'die Hunde'). Ingest-time, so the app never mutates."""
    text = value.strip()
    if not text:
        return text
    head, sep, tail = text.partition(" ")
    if sep and head.lower() in _NOUN_ARTICLES and tail:
        return f"{head.lower()} {tail[:1].upper()}{tail[1:]}"
    return text[:1].upper() + text[1:]


def compute_content_hash(row: dict[str, Any]) -> str:
    concatenated = "|".join(
        str(row[k]) if row[k] is not None else ""
        for k in sorted(row.keys())
        if k not in ("id", "content_hash")
    )
    return hashlib.sha256(concatenated.encode()).hexdigest()


# ---------------------------------------------------------------------------
# Header resolution
# ---------------------------------------------------------------------------

def _resolve_headers(table: Table, actual: list[str], filename: str) -> dict[str, int]:
    """Map every readable field to its sheet column index.

    Returns {"core:<field>": idx, "lang:<code>:<field>": idx, "special:<name>": idx}.
    Canonical-vs-alias: the FIRST header name (canonical, then aliases in order)
    present in the sheet wins; a sheet carrying BOTH spellings of one field is an
    error (the ingest cannot know which column to trust — same rule as v1
    duplicates). Unknown non-tooling columns warn loudly (a typo'd language
    column must never vanish silently)."""
    index: dict[str, int] = {}
    problems: list[str] = []

    def place(key: str, names: tuple[str, ...], required: bool) -> None:
        present = [n for n in names if n in actual]
        if len(present) > 1:
            problems.append(
                f"  both {present!r} present — one field, one column (aliases of each other)"
            )
            return
        if not present:
            if required:
                problems.append(f"  missing required column: {names[0]!r} (aliases: {list(names[1:])})")
            return
        name = present[0]
        if actual.count(name) > 1:
            problems.append(f"  duplicated column {name!r} — remove the duplicate")
            return
        index[key] = actual.index(name)

    for fieldname, names in table.columns.items():
        required = fieldname in ("level", "type", "word", "german_sentence") or (
            fieldname in table.required and fieldname != "sense"
        )
        # sense is always optional
        place(f"core:{fieldname}", names, required and fieldname != "sense")

    for code, lang in LANGUAGES.items():
        for fieldname, names in lang.columns.items():
            place(f"lang:{code}:{fieldname}", names, required=False)

    for name in registry.SPECIAL_HEADERS:
        place(f"special:{name}", (name,), required=False)

    # English is the required language: its word+sentence columns must exist.
    for fieldname in ("word", "sentence"):
        if f"lang:{REQUIRED_LANGUAGE}:{fieldname}" not in index:
            names = LANGUAGES[REQUIRED_LANGUAGE].columns[fieldname]
            problems.append(f"  missing required column: {names[0]!r} (aliases: {list(names[1:])})")

    if problems:
        raise ValidationError(f"'{filename}' column problems:\n" + "\n".join(problems))

    known = known_headers_for(table)
    unknown = [c for c in actual if c and c not in known and not is_ignored_header(c)]
    if unknown:
        logger.warning(
            "'%s': %d unrecognised column(s) — IGNORED. If one of these is a language "
            "column, fix its header or add it to the registry: %s",
            filename, len(unknown), unknown,
        )
    return index


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

def read_dataset(table_name: str, skip_invalid: bool = False) -> TableDataset:
    table = TABLES[table_name]
    path = DATA_DIR / table.xlsx
    if not path.exists():
        raise ValidationError(f"'{table.xlsx}' not found at {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if len(wb.sheetnames) > 1:
            raise ValidationError(
                f"'{table.xlsx}' has multiple sheets: {wb.sheetnames} — remove all but one."
            )
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValidationError(f"'{table.xlsx}' is completely empty.")

        actual = [
            str(c).replace("\xa0", " ").strip() if c is not None else ""
            for c in header_row
        ]
        while actual and actual[-1] == "":
            actual.pop()

        index = _resolve_headers(table, actual, table.xlsx)

        def cell(raw_row: tuple, key: str) -> str | None:
            i = index.get(key, -1)
            return _clean(raw_row[i]) if 0 <= i < len(raw_row) else None

        ds = TableDataset()
        errors: list[str] = []
        # v2 id -> (row_num, word, sense) for homonym-collision reporting
        seen: dict[str, tuple[int, str, str]] = {}

        for row_num, raw_row in enumerate(rows_iter, start=2):
            if all(_clean(c) is None for c in raw_row):
                continue

            row_errors: list[str] = []
            level = cell(raw_row, "core:level") or ""
            word = cell(raw_row, "core:word")
            label = repr(word) if word else "(unknown word)"

            if not level:
                row_errors.append(f"  Row {row_num}: 'Level' is empty")
            elif not VALID_LEVEL.match(level):
                row_errors.append(
                    f"  Row {row_num}: invalid Level {level!r} "
                    f"(A1/A2/B1/B2/C1/C2, optionally .1 or .2)"
                )
            if not word:
                row_errors.append(f"  Row {row_num}: word column is empty")

            core: dict[str, Any] = {}
            for fieldname in table.columns:
                core[fieldname] = cell(raw_row, f"core:{fieldname}")

            for fieldname in table.required:
                if not core.get(fieldname):
                    row_errors.append(
                        f"  Row {row_num}: required field '{fieldname}' is empty (word={label})"
                    )

            type_raw = core.get("type")
            type_norm = str(type_raw).strip().lower() if type_raw else ""
            if type_norm and type_norm not in table.allowed_types:
                allowed = "/".join(sorted(table.allowed_types))
                row_errors.append(
                    f"  Row {row_num}: invalid Type {type_raw!r} — "
                    f"'{table_name}' allows only {allowed} (word={label})"
                )

            if table_name == "nouns" and core.get("article"):
                # Single article, or a slash combination for nominalized forms /
                # multi-gender nouns ("der/die Mitarbeitende", "der/die/das Joghurt").
                art = str(core["article"]).strip().lower()
                parts = art.split("/")
                if (any(p not in GERMAN_ARTICLES for p in parts)
                        or len(set(parts)) != len(parts)):
                    row_errors.append(
                        f"  Row {row_num}: invalid Article {core['article']!r} — "
                        f"der/die/das or a slash combination without repeats (word={label})"
                    )

            sense = core.get("sense") or ""
            if sense:
                sense = registry.normalize_for_id(sense)
                if not VALID_SENSE.match(sense):
                    row_errors.append(
                        f"  Row {row_num}: invalid Sense {core['sense']!r} — "
                        f"lowercase letters/digits/hyphen, max 24 chars (word={label})"
                    )
            core["sense"] = sense or None

            # The required language (English) must be present, like v1.
            en_word = cell(raw_row, f"lang:{REQUIRED_LANGUAGE}:word")
            en_sentence = cell(raw_row, f"lang:{REQUIRED_LANGUAGE}:sentence")
            if not en_word:
                row_errors.append(f"  Row {row_num}: English word is empty (word={label})")
            if not en_sentence:
                row_errors.append(f"  Row {row_num}: English sentence is empty (word={label})")

            if row_errors:
                errors.extend(row_errors)
                ds.skipped += 1
                if word and type_norm:
                    article = str(core.get("article") or "")
                    ds.protected.add(compute_word_id(type_norm, word, article, sense))
                if word and level and VALID_LEVEL.match(level):
                    ds.protected_aliases.add(compute_legacy_id(level, str(word)))
                continue

            article = str(core.get("article") or "")
            word_id = compute_word_id(type_norm, str(word), article, sense)
            prev = seen.get(word_id)
            if prev is not None:
                errors.append(
                    f"  Rows {prev[0]} and {row_num}: {word!r} (type={type_norm}"
                    f"{', article=' + article if article else ''}) collapse to ONE identity — "
                    f"if these are different senses (homonyms), give each a Sense tag "
                    f"(current: {prev[2] or '(none)'} / {sense or '(none)'}); "
                    f"if they duplicate, remove one."
                )
                ds.skipped += 1
                continue
            seen[word_id] = (row_num, str(word), sense)

            # Normalize once at ingest (v1 doctrine): lowercase type/article,
            # capitalized German nouns and plurals.
            core["type"] = type_norm
            core["level"] = level
            if table_name == "nouns":
                if core.get("article"):
                    core["article"] = str(core["article"]).strip().lower()
                if core.get("word"):
                    core["word"] = _capitalize_noun(str(core["word"]))
                if core.get("plural"):
                    core["plural"] = _capitalize_noun(str(core["plural"]))
                core["image"] = _to_bool01(
                    raw_row[index["special:Image"]]
                    if "special:Image" in index and index["special:Image"] < len(raw_row)
                    else None
                )
            core["free"] = _to_bool01(
                raw_row[index["special:Free"]]
                if "special:Free" in index and index["special:Free"] < len(raw_row)
                else None
            )
            core["id"] = word_id
            core["content_hash"] = compute_content_hash(core)
            ds.core.append(core)

            # Alias row: the v1 identity of this exact row (level|word), for the
            # re-key migration and media re-label (WD-ID-4/5).
            legacy = compute_legacy_id(level, str(word))
            alias = {"id": legacy, "new_id": word_id, "reason": "v2-rekey"}
            alias["content_hash"] = compute_content_hash(alias)
            ds.aliases.append(alias)

            # Translation rows: one per language with content (word present).
            for code, lang in LANGUAGES.items():
                values = {
                    f: cell(raw_row, f"lang:{code}:{f}")
                    for f in TRANSLATION_FIELDS
                    if f"lang:{code}:{f}" in index
                }
                if not values.get("word"):
                    continue
                trow: dict[str, Any] = {
                    "id": f"{word_id}:{code}",
                    "word_id": word_id,
                    "lang": code,
                }
                for f in TRANSLATION_FIELDS:
                    trow[f] = values.get(f)
                trow["content_hash"] = compute_content_hash(trow)
                ds.translations.append(trow)
                ds.coverage[code] = ds.coverage.get(code, 0) + 1

        if errors:
            if not skip_invalid:
                raise ValidationError(
                    f"'{table.xlsx}' has {len(errors)} validation error(s):\n"
                    + "\n".join(errors)
                    + "\n  (re-run with --skip-invalid to skip these rows and sync the rest)"
                )
            logger.warning(
                "'%s': skipping %d invalid row(s) (--skip-invalid):\n%s",
                table.xlsx, ds.skipped, "\n".join(errors),
            )

        if not ds.core:
            raise ValidationError(f"'{table.xlsx}' contains no valid data rows.")
        return ds
    finally:
        wb.close()


def coverage_report(datasets: dict[str, TableDataset]) -> list[str]:
    """Per-language coverage lines for the publish preview (LG-FR-14). A variant
    counts its own rows only — its BASE fallback covers the rest by design."""
    total = sum(len(d.core) for d in datasets.values())
    lines: list[str] = []
    for code, lang in LANGUAGES.items():
        n = sum(d.coverage.get(code, 0) for d in datasets.values())
        if code == REQUIRED_LANGUAGE:
            lines.append(f"  {lang.display} ({code}): {n}/{total} (required)")
        elif n:
            base = f", falls back to {lang.base}" if lang.base else ""
            lines.append(f"  {lang.display} ({code}): {n}/{total}{base}")
    return lines
