"""
Image GENERATION — orchestrate the noun-image pipeline: source, review-queue, approve into the
decisions store + the durable R2 master mirror. Publishing packs/manifests is media_publish.py's
job (it reads the decisions store). The old pack-publish path here rewrote the shared manifest with
name-keyed metas and could clobber media_publish's output (audit 2026-07-19, H7); it was removed.

Mirrors audio_sync.py: collect the flagged nouns, do only the work that's needed (idempotent via the
decisions store) — source each flagged noun's image and queue it for manual approval.

Per noun (unless --no-source), the engine GENERATES the image(s) via Azure Foundry and queues EVERY
image for manual review (no stock search, no auto-verify) — a person approves each in image_review.py.
Approved images are written to the local master cache AND mirrored to R2 (image/files/<hash>.heic) so a
fresh checkout never re-generates. Re-runs skip everything already settled for its current content.

Usage:
  python image_sync.py                 # source what's needed into decisions + master mirror
  python image_sync.py --dry-run       # build/report locally, upload nothing
  python image_sync.py --no-source     # skip sourcing (reconcile/prune only)
  python image_sync.py --limit 50      # process at most 50 not-yet-settled nouns this run
  python image_sync.py --workers 4     # generate 4 nouns in parallel (errors retry, never settle)
  python image_sync.py --include-unflagged --limit 100  # also do nouns NOT marked x/y (big set — cap it)
  python image_sync.py --free-first    # generate the free-tier nouns before the rest
  python image_sync.py --prune-files   # delete orphan image/files masters in R2
  python image_sync.py --delete-all    # RESET images to a clean slate (R2 images + local state; audio untouched)

Keys (sync/.env): R2_*, PIXABAY_API_KEY, PEXELS_API_KEY, AZURE_FOUNDRY_* (see image_config.py).
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import shutil
import sys
from datetime import date
from pathlib import Path
from typing import Any

from dotenv import load_dotenv

import image_config as cfg
import image_decisions
import image_engine
import media_delivery
import sync  # read_excel / TABLE_CONFIG / logging setup / ValidationError

load_dotenv(Path(__file__).parent / ".env")

logger = logging.getLogger("image_sync")

SAVE_EVERY = 20  # checkpoint decisions+queue this often so a crash never loses sourcing/review work


# ---------------------------------------------------------------------------
# Inputs + local master cache
# ---------------------------------------------------------------------------

def collect_nouns(include_unflagged: bool = False) -> list[dict[str, Any]]:
    """The nouns eligible for an image. By default only those flagged in the sheet (Image column =
    x or y → row['image'] == 1). With include_unflagged=True, every noun is returned — used both by
    `--include-unflagged` sourcing and, always, for packing/pruning so an approved unflagged noun is
    shipped and never pruned on a marked-only run."""
    rows, _, _ = sync.read_excel("nouns")
    if include_unflagged:
        return rows
    return [r for r in rows if r.get("image") == 1]


def _master_path(content_hash: str) -> Path:
    return cfg.CACHE_DIR / f"{content_hash}.{cfg.FILE_EXT}"


def _write_master(webp: bytes, content_hash: str) -> None:
    cfg.CACHE_DIR.mkdir(parents=True, exist_ok=True)
    _master_path(content_hash).write_bytes(webp)


def _ensure_master_local(client, bucket: str | None, content_hash: str) -> bool:
    """Make sure the processed WebP master is on disk, pulling it from R2 if needed (fresh checkout).
    Returns True if available locally afterwards."""
    if _master_path(content_hash).exists():
        return True
    if client is None:
        return False
    return media_delivery.download_file(client, bucket or "", cfg.FILES_PREFIX, content_hash,
                                        cfg.FILE_EXT, _master_path(content_hash))


# ---------------------------------------------------------------------------
# Review queue (IO lives in image_decisions — one implementation for all tools)
# ---------------------------------------------------------------------------

_load_review_queue = image_decisions.load_review_queue
_save_review_queue = image_decisions.save_review_queue


def _queue_entry(noun: dict[str, Any], outcome) -> dict[str, Any]:
    """A review-queue record: the noun context + each candidate's provenance and its local preview
    (image_cache/<hash>.heic), so image_review.py can show the FINAL framing and write back a pick."""
    return {
        "word": noun.get("word"),
        "article": noun.get("article"),
        "english": noun.get("english"),
        "german_sentence": noun.get("german_sentence"),
        "english_sentence": noun.get("english_sentence"),  # the sentence fed to image generation
        "candidates": [{
            "content_hash": pc.content_hash,
            "source": pc.candidate.source,
            "source_id": pc.candidate.source_id,
            "url": pc.candidate.page_url or pc.candidate.image_url,
            "license": pc.candidate.license,
            "kind": pc.kind,
            "verifier": pc.verifier,
            "clip": pc.clip,
        } for pc in outcome.candidates],
    }


# ---------------------------------------------------------------------------
# Sourcing pass
# ---------------------------------------------------------------------------

_LEVEL_RE = re.compile(r"^([abc])([12])(?:\.([12]))?$")


def _level_rank(level: str) -> tuple[int, int, int]:
    """Sort key so generation runs in CEFR order: A1.1, A1.2, A2.1, … C2.2. Unknown levels sort last."""
    m = _LEVEL_RE.match((level or "").strip().lower())
    if not m:
        return (9, 9, 9)
    return ("abc".index(m.group(1)), int(m.group(2)), int(m.group(3) or 0))


def _source_pass(nouns, store, queue, opts, *, client, bucket, dry_run, limit, free_first=False,
                 workers=1) -> None:
    def _should_generate(n: dict[str, Any]) -> bool:
        # (Re)generate when the content changed or is new (approved/none/review all re-trigger on a
        # changed fingerprint). A noun already settled for its current content is NOT regenerated —
        # except a `review` noun whose queued candidates were lost, which we repopulate.
        rec = store.get(n["id"]) or {}
        if rec.get("replace_requested"):
            # Zero-gap replacement: candidates already generated for the CURRENT content and
            # waiting in the review queue count as done for this round; anything else
            # (not queued yet, queue lost, content edited since) generates a fresh round.
            return not (n["id"] in queue
                        and rec.get("replace_fingerprint") == image_decisions.input_fingerprint(n))
        if image_decisions.needs_processing(store, n):
            return True
        return rec.get("status") == "review" and n["id"] not in queue

    todo = [n for n in nouns if _should_generate(n)]

    # Ordering:
    #  - with --free-first: ALL free-tier nouns come first — no free word is left before any non-free
    #    word. Within each tier, reviewer-noted ones first, then CEFR level ascending.
    #  - otherwise: reviewer-noted first, then CEFR level ascending (A1.1 → C2.2).
    # Stable sort → original sheet order preserved within a group.
    def _order(n: dict[str, Any]) -> tuple:
        noted = 0 if image_decisions.get_note(opts, n["id"]) else 1
        level = _level_rank(n.get("level", ""))
        if free_first:
            return (0 if n.get("free") else 1, noted, level)   # free is the ABSOLUTE first key
        return (noted, level)
    todo.sort(key=_order)

    if limit:
        todo = todo[:limit]
    n_noted = sum(1 for n in todo if image_decisions.get_note(opts, n["id"]))
    n_free = sum(1 for n in todo if n.get("free"))
    if free_first:
        order_desc = f"ALL free-tier first ({n_free} free), then noted, then CEFR level"
    else:
        order_desc = "noted first, then CEFR level (A1.1→C2.2)" + (f" ({n_noted} noted)" if n_noted else "")
    logger.info("Sourcing: %d of %d flagged noun(s) need work — %s.", len(todo), len(nouns), order_desc)
    if dry_run:
        logger.info("[DRY RUN] Skipping sourcing (would process %d).", len(todo))
        return

    today = date.today().isoformat()
    total = len(todo)
    stats = {"approved": 0, "review": 0, "none": 0, "error": 0}

    def _generate(noun: dict[str, Any]) -> "image_engine.Outcome":
        """Pure, side-effect-free: the slow FLUX call(s) + crop/encode. Safe to run on a worker
        thread — touches no shared state and never writes the store/queue/R2."""
        return image_engine.process_noun(
            noun,
            use_sentence=image_decisions.uses_sentence(opts, noun["id"]),
            note=image_decisions.get_note(opts, noun["id"]),
        )

    def _apply(i: int, noun: dict[str, Any], outcome) -> None:
        """Reduce step — runs ONLY on the main thread, so store/queue/stats need no locking.
        An 'error' outcome writes NOTHING (the noun stays unsettled and is retried next run)."""
        nid = noun["id"]
        lvl = (noun.get("level") or "?").upper()
        tier = "free" if noun.get("free") else "paid"
        src = "review+prompt" if image_decisions.get_note(opts, nid) else "new"
        res = f"review×{len(outcome.candidates)}" if outcome.status == "review" else outcome.status
        logger.info("  [%d/%d] %-5s %-4s %-13s %s (%s) → %s",
                    i, total, lvl, tier, src, noun.get("word", ""), noun.get("english", ""), res)
        if outcome.status == "approved":
            pc = outcome.chosen
            _write_master(pc.master, pc.content_hash)
            media_delivery.upload_file(client, bucket, cfg.FILES_PREFIX, pc.content_hash, cfg.FILE_EXT,
                                       _master_path(pc.content_hash))
            image_decisions.record_approved(
                store, noun, source=pc.candidate.source, source_id=pc.candidate.source_id,
                url=pc.candidate.page_url or pc.candidate.image_url, license=pc.candidate.license,
                kind=pc.kind, content_hash=pc.content_hash, approved_by="auto", today=today,
                verifier=pc.verifier,
            )
            queue.pop(nid, None)
        elif outcome.status == "review":
            for pc in outcome.candidates:
                _write_master(pc.master, pc.content_hash)  # local masters; review server transcodes to JPEG
            rec = store.get(nid) or {}
            if rec.get("status") == "approved" and rec.get("replace_requested"):
                # Zero-gap replacement: KEEP the approved record (the current image keeps
                # shipping) while the fresh candidates await review. Stamp the fingerprint
                # the candidates were generated for so a later content edit re-triggers.
                rec["replace_fingerprint"] = image_decisions.input_fingerprint(noun)
                rec["updated"] = today
            else:
                image_decisions.mark_review(store, noun, today)
            queue[nid] = _queue_entry(noun, outcome)
        elif outcome.status == "error":
            # Generation FAILED — do NOT settle. Leave the decision untouched so the next run's
            # _should_generate() picks it up again. (No queue/store mutation here on purpose.)
            pass
        else:  # "none" — a clean miss (e.g. content-safety blocked): settle so we don't loop forever.
            rec = store.get(nid) or {}
            if rec.get("status") == "approved" and rec.get("replace_requested"):
                # Replacement round produced nothing usable — keep the existing image rather
                # than shipping a blank. Clear the request so we don't loop forever; asking
                # again (media_replace row or reviewer note) starts a fresh round.
                rec.pop("replace_requested", None)
                rec.pop("replace_fingerprint", None)
                rec["updated"] = today
                queue.pop(nid, None)
                logger.info("  replacement for %s produced nothing — keeping the current image",
                            noun.get("word", nid))
            else:
                image_decisions.mark_none(store, noun, today)
                queue.pop(nid, None)
        stats[outcome.status] = stats.get(outcome.status, 0) + 1

        if i % SAVE_EVERY == 0:
            image_decisions.save(store)
            _save_review_queue(queue)
            logger.info("  …%d/%d (approved=%d review=%d none=%d error=%d)", i, total,
                        stats["approved"], stats["review"], stats["none"], stats["error"])

    if workers and workers > 1:
        # Parallel generate (threads) → serial apply (main thread). process_noun is pure; only the
        # cheap reduce mutates shared state, so no locks are needed and nouns never confuse each other.
        from concurrent.futures import ThreadPoolExecutor, as_completed
        logger.info("Generating with %d parallel worker(s).", workers)
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futs = {ex.submit(_generate, n): n for n in todo}
            for i, fut in enumerate(as_completed(futs), 1):
                noun = futs[fut]
                try:
                    outcome = fut.result()
                except Exception as exc:  # noqa: BLE001 — anything unexpected: count as error, never settle
                    stats["error"] += 1
                    logger.warning("  [%d/%d] generation FAILED for %s (%s) — left unsettled, will retry",
                                   i, total, noun.get("word", ""), exc)
                    continue
                _apply(i, noun, outcome)
    else:
        for i, noun in enumerate(todo, 1):
            _apply(i, noun, _generate(noun))

    image_decisions.save(store)
    _save_review_queue(queue)
    logger.info("Sourcing done: approved=%d, review=%d, none=%d, error=%d (errors retry next run).",
                stats["approved"], stats["review"], stats["none"], stats["error"])
    return stats


# ---------------------------------------------------------------------------
# Pack build + publish
# ---------------------------------------------------------------------------

def _flag_ids_in_sheet(ids: set[str]) -> int:
    """Set nouns.xlsx Image → 'y' for the given noun ids (identity = compute_id(level, word)), when a
    row is not already flagged. Rewrites the sheet once; returns the count newly flagged. Safe no-op
    when nothing needs flagging or the file is missing/locked."""
    if not ids:
        return 0
    import openpyxl
    from openpyxl.styles import PatternFill
    filename, _ = sync.TABLE_CONFIG["nouns"]
    path = sync.DATA_DIR / filename
    if not path.exists():
        logger.warning("Cannot flag words in the sheet — %s not found.", path)
        return 0
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as exc:  # noqa: BLE001 — e.g. the file is open in Excel/Numbers
        logger.warning("Cannot open %s to flag words (%s) — leave it closed and re-run.", filename, exc)
        return 0
    try:
        ws = wb.active
        red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # highlight changed cells
        header = [(str(c.value).replace("\xa0", " ").strip() if c.value is not None else "") for c in ws[1]]
        cols = {n: (header.index(n) + 1 if n in header else -1) for n in ("Level", "Word", "Image")}
        if min(cols.values()) < 0:
            logger.warning("Cannot flag words — missing Level/Word/Image column in %s.", filename)
            return 0
        remaining = set(ids)
        flagged = 0
        for r in range(2, ws.max_row + 1):
            if not remaining:
                break
            lvl = sync._clean(ws.cell(r, cols["Level"]).value)
            word = sync._clean(ws.cell(r, cols["Word"]).value)
            if lvl is None or word is None:
                continue
            if sync.compute_id(str(lvl), str(word)) not in remaining:
                continue
            remaining.discard(sync.compute_id(str(lvl), str(word)))
            cell = ws.cell(r, cols["Image"])
            cur_norm = str(cell.value).replace("\xa0", " ").strip().lower() if cell.value is not None else ""
            if cur_norm in sync._TRUTHY:
                continue                                    # already flagged — leave as-is
            cell.value = "y"
            cell.fill = red                                 # mark the changed cell red
            flagged += 1
        if flagged:
            wb.save(path)
        return flagged
    finally:
        wb.close()


def flag_approved_in_sheet(store) -> int:
    """Reconcile the sheet with reality: every noun that has an APPROVED image gets Image='y' in
    nouns.xlsx (the source of truth for the app's image flag). Idempotent — only unflagged rows are
    written. Run `python sync.py` afterwards to propagate the new flags to D1. Returns count flagged."""
    return _flag_ids_in_sheet(set(image_decisions.approved(store)))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _reencode_pass(store, *, client, bucket, dry_run) -> None:
    """Migrate already-approved masters to Apple-conformant HEIC (sips) WITHOUT re-sourcing or
    regenerating — fixes the 'Invalid value for reserved bit' warning for legacy x265 images. Each
    approved master is transcoded; its content_hash changes, so the decision is updated and the new
    master uploaded (the old one is pruned afterwards by run()). Idempotent: skips masters already
    marked `encoder: "sips"`, so re-running does nothing."""
    approved = image_decisions.approved(store)
    todo = [(nid, rec) for nid, rec in approved.items() if rec.get("encoder") != "sips"]
    logger.info("Re-encode: %d of %d approved master(s) need Apple-conformant HEIC.", len(todo), len(approved))
    if dry_run:
        logger.info("[DRY RUN] Would re-encode %d master(s); nothing written.", len(todo))
        return
    done = 0
    for nid, rec in todo:
        old_hash = rec.get("content_hash")
        if not old_hash or not _ensure_master_local(client, bucket, old_hash):
            logger.warning("  master %s missing locally/R2 — skipping %s", (old_hash or "?")[:8], nid)
            continue
        try:
            new_bytes = image_engine.reencode_master(_master_path(old_hash).read_bytes())
        except Exception as exc:  # noqa: BLE001 — log and continue; one bad file shouldn't abort the batch
            logger.warning("  re-encode failed for %s: %s", nid, exc)
            continue
        new_hash = image_engine._content_hash(new_bytes)
        _write_master(new_bytes, new_hash)
        media_delivery.upload_file(client, bucket, cfg.FILES_PREFIX, new_hash, cfg.FILE_EXT, _master_path(new_hash))
        rec["content_hash"] = new_hash
        rec["encoder"] = "sips"
        done += 1
        if done % SAVE_EVERY == 0:
            image_decisions.save(store)
            logger.info("  …%d/%d re-encoded", done, len(todo))
    image_decisions.save(store)
    logger.info("Re-encoded %d master(s) to Apple-conformant HEIC.", done)


def run(*, dry_run: bool, no_source: bool, limit: int, prune_files: bool,
        free_first: bool = False, reencode_masters: bool = False, workers: int = 1,
        include_unflagged: bool = False, client=None, bucket: str | None = None) -> None:
    logger.info("Reading nouns…")
    all_nouns = collect_nouns(include_unflagged=True)   # every noun in the sheet — for packing + prune liveness
    flagged = [n for n in all_nouns if n.get("image") == 1]
    source_nouns = all_nouns if include_unflagged else flagged
    logger.info("Flagged for images: %d noun(s)%s.", len(flagged),
                f"; sourcing all {len(source_nouns)} (incl. unflagged)" if include_unflagged else "")

    store = image_decisions.load()
    queue = _load_review_queue()
    opts = image_decisions.load_prompt_opts()

    srcstats = None
    if reencode_masters:
        _reencode_pass(store, client=client, bucket=bucket, dry_run=dry_run)
    elif not no_source:
        srcstats = _source_pass(source_nouns, store, queue, opts, client=client, bucket=bucket,
                                dry_run=dry_run, limit=limit, free_first=free_first, workers=workers)
    else:
        logger.info("Skipping sourcing (--no-source); building from existing decisions + cache.")

    # Prune decisions/queue ONLY for nouns removed from the sheet entirely — an approved noun (flagged
    # or not) stays, so packing below ships it. (Uses all_nouns, so de-flagging never drops an image.)
    live_ids = {n["id"] for n in all_nouns}
    removed = image_decisions.prune(store, live_ids)
    if removed:
        logger.info("Pruned %d decision(s) for noun(s) removed from the sheet.", removed)
    for nid in [q for q in queue if q not in live_ids]:
        queue.pop(nid, None)

    # Pack/manifest publishing is media_publish.py's job (it reads image_decisions + the
    # master cache to build the immutable v2 packs/catalogs). This tool only SOURCES images
    # into decisions + the master mirror. The old pack-publish here rewrote the shared
    # manifest with name-keyed metas and could clobber media_publish's output (audit H7).

    # Re-encoding replaces masters (new hashes), so the old ones are now orphaned → always prune then.
    if (prune_files or reencode_masters) and not dry_run and client is not None:
        media_delivery.prune_orphan_files(client, bucket, cfg.FILES_PREFIX, cfg.FILE_EXT,
                                          image_decisions.live_content_hashes(store))

    image_decisions.save(store)
    _save_review_queue(queue)

    # Reconcile the source sheet: any approved (incl. previously-unflagged) noun gets Image='y' so the
    # app will show it. sync.py then propagates the flag to D1.
    if not dry_run:
        newly = flag_approved_in_sheet(store)
        if newly:
            logger.info("Flagged %d newly-approved word(s) as 'y' (red) in nouns.xlsx — run `python sync.py` "
                        "to propagate the image flag to D1.", newly)

    # Queue length, not store status: a zero-gap replacement awaits review while its
    # store record stays "approved" (the old image keeps shipping).
    n_review = len(queue)
    if n_review:
        logger.info("%d noun(s) await review — run: python image_review.py", n_review)

    # End-of-run "go live" summary for --include-unflagged: what still needs review + a sync to reach D1.
    if include_unflagged and srcstats is not None:
        generated = srcstats["review"] + srcstats["approved"]
        logger.info("─ Unflagged-words run summary ─")
        logger.info("  generated this run : %d (%d queued for review, %d errored → retry next run)",
                    generated, srcstats["review"], srcstats["error"])
        logger.info("  to make them live  :")
        if n_review:
            logger.info("    1) python image_review.py   # pick images (auto-flags them 'y', red, in nouns.xlsx)")
            logger.info("    2) python sync.py           # push the new image flags to D1 so the app shows them")
        else:
            logger.info("    python sync.py              # push the new image flags to D1 so the app shows them")
    logger.info("Image sync complete.")


def _confirm_reset() -> bool:
    """Guard the destructive reset: proceed only if the operator types 'delete images'."""
    print("\n[CONFIRM] This DELETES every image from R2 (all image/* packs + image/files masters),")
    print("          rewrites the shared manifest WITHOUT the image packs (audio is left untouched),")
    print("          and removes the local image_cache/, review queue, prompt notes and decisions.")
    print("          Audio is NOT affected. This cannot be undone.")
    try:
        return input("Type 'delete images' to proceed (anything else aborts): ").strip().lower() == "delete images"
    except (EOFError, KeyboardInterrupt):
        print()
        return False


def reset_all(*, client, bucket: str | None, dry_run: bool) -> None:
    """Wipe the IMAGE side back to a clean slate — images only, never audio.

    R2 (if online): delete every image/files master. (Pack/manifest state is media_publish.py's
    domain — the next publish rebuilds the image packs/catalog from the now-empty decisions and
    a promote drops them from the channel; this tool no longer touches the shared manifest.)
    Locally: remove the master cache, the review directory (queue + previews), the prompt-notes
    file and the decisions store.
    """
    # 1. R2: delete every image master (audio untouched). Packs/catalogs are rebuilt by
    #    media_publish from the emptied decisions store; gc reclaims the orphaned pack objects.
    if not dry_run and client is not None:
        logger.info("R2: deleting all image/files masters…")
        media_delivery.prune_orphan_files(client, bucket, cfg.FILES_PREFIX, cfg.FILE_EXT, set())
    else:
        logger.info("[DRY RUN] would delete every image/files master in R2.")

    # 2. Local: cache, review dir (queue + previews), prompt notes, decisions.
    targets = [cfg.CACHE_DIR, cfg.REVIEW_DIR, image_decisions.PROMPT_OPTS_PATH, cfg.DECISIONS_PATH]
    for path in targets:
        if not path.exists():
            continue
        if dry_run:
            logger.info("[DRY RUN] would delete %s", path.name)
            continue
        if path.is_dir():
            shutil.rmtree(path)
        else:
            path.unlink()
        logger.info("Deleted %s", path.name)
    logger.info("Image reset complete — clean slate." + (" (dry run — nothing changed)" if dry_run else ""))


def main() -> None:
    parser = argparse.ArgumentParser(description="Source, verify and sync noun images to Cloudflare R2.")
    parser.add_argument("--dry-run", action="store_true", help="Build/report locally; upload nothing.")
    parser.add_argument("--no-source", action="store_true", help="Skip sourcing; build/publish from decisions+cache.")
    parser.add_argument("--limit", type=int, default=0, help="Process at most N not-yet-settled nouns this run.")
    parser.add_argument("--free-first", action="store_true",
                        help="Generate the free-tier (Free=1) nouns before the rest (composes with --limit).")
    parser.add_argument("--include-unflagged", action="store_true",
                        help="Also source images for nouns NOT marked x/y in the sheet (default: marked only). "
                             "Big set — pair with --limit/--workers to control cost. Approved ones are packed + shipped.")
    parser.add_argument("--workers", type=int, default=1, metavar="N",
                        help="Generate N nouns in parallel (threads). Default 1. 3–4 is a good balance "
                             "vs. Azure rate limits; failures are isolated and retried, never settled.")
    parser.add_argument("--reencode-masters", action="store_true",
                        help="Transcode all approved masters to Apple-conformant HEIC (sips) without "
                             "re-sourcing/regenerating — fixes the HEVC 'reserved bit' warning. One-shot, idempotent.")
    parser.add_argument("--prune-files", action="store_true",
                        help="After publishing, delete image/files masters in R2 no longer referenced.")
    parser.add_argument("--delete-all", action="store_true",
                        help="Reset images to a clean slate: delete ALL images from R2 (packs + masters, "
                             "rewriting the shared manifest without them — audio untouched) and remove the "
                             "local cache, review queue, prompt notes and decisions. Prompts to confirm.")
    parser.add_argument("--yes", action="store_true", help="Skip the --delete-all confirmation prompt.")
    verbosity = parser.add_mutually_exclusive_group()
    verbosity.add_argument("-v", "--verbose", action="store_true")
    verbosity.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args()

    sync._setup_logging(args.verbose, args.quiet)
    args.workers = max(1, min(args.workers, 8))   # clamp to a sane range

    if not args.dry_run:
        required = ["R2_ACCOUNT_ID", "R2_ACCESS_KEY_ID", "R2_SECRET_ACCESS_KEY", "R2_BUCKET"]
        missing = [k for k in required if not os.environ.get(k)]
        if missing:
            logger.error("Missing environment variables: %s", ", ".join(missing))
            logger.error("Set them in sync/.env (or run with --dry-run).")
            sys.exit(1)

    client = None
    bucket = os.environ.get("R2_BUCKET")
    if not args.dry_run:
        client = media_delivery.r2_client()

    if args.delete_all:
        if not args.dry_run and not args.yes and not _confirm_reset():
            logger.info("Aborted — nothing deleted.")
            sys.exit(0)
        reset_all(client=client, bucket=bucket, dry_run=args.dry_run)
        return

    try:
        run(dry_run=args.dry_run, no_source=args.no_source, limit=args.limit,
            prune_files=args.prune_files, free_first=args.free_first,
            reencode_masters=args.reencode_masters, workers=args.workers,
            include_unflagged=args.include_unflagged, client=client, bucket=bucket)
    except sync.ValidationError as exc:
        logger.error("Validation failed: %s", exc)
        sys.exit(1)


if __name__ == "__main__":
    main()
