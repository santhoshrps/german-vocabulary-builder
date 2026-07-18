"""Words publish CLI (v2): spreadsheet -> per-environment content database.

V2 flow (WD-ID / LG-FR-9..14): sync/dataset.py reads the sheets registry-driven
(sync/registry.py) into German core rows + translations rows + id aliases, and this
CLI diffs and publishes all five tables against the target environment's write
worker (sync/envs.py; --env, default dev, typed gate for prod).

LEGACY-COMPAT surface (remove at P2): TABLE_CONFIG / read_excel / compute_id are
the v1 reader, kept ONLY because the media pipeline (audio_sync, image_sync,
media_replace, image_regen/review) still keys its caches and R2 objects by v1 ids
until the P2 media re-label. New code imports dataset/registry, never these.
"""

import argparse
import hashlib
import hmac
import logging
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

import httpx
import openpyxl
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

import dataset as dataset_mod
import envs
from registry import TABLES as V2_TABLES

logger = logging.getLogger("sync")

# Set by _load_environment() (v2 flow); the module-level defaults keep the LEGACY
# read path importable for the media pipeline without an environment.
WORKER_URL = os.environ.get("WORKER_URL", "").rstrip("/")
API_KEY = os.environ.get("API_KEY", "")
DATA_DIR = Path(__file__).parent.parent / "data"

# ---------------------------------------------------------------------------
# LEGACY-COMPAT (v1 reader) — media pipeline only; remove at the P2 re-label.
# ---------------------------------------------------------------------------

TABLE_CONFIG: dict[str, tuple[str, list[str]]] = {
    "verbs": (
        "verbs.xlsx",
        [
            "Level", "Capital", "Type", "German_Word", "English_Word",
            "German_Sentence", "English_Sentence",
            "ich", "du", "er_sie_es", "wir", "ihr", "sie_Sie",
            "past_participle", "simple_past", "Free",
        ],
    ),
    "nouns": (
        "nouns.xlsx",
        [
            "Level", "Capital", "Type", "German_Article", "German_Word", "German_Plural",
            "Image", "English_Word", "German_Sentence", "English_Sentence", "Free",
        ],
    ),
    "adverbs_adjectives": (
        "adverbs_adjectives.xlsx",
        [
            "Level", "Capital", "Type", "German_Word", "English_Word",
            "German_Sentence", "English_Sentence", "German_Comparative", "German_Superlative", "Free",
        ],
    ),
}

HEADER_TO_DB: dict[str, str] = {
    "sie_Sie": "sie_sie",
    "German_Word": "word",
    "German_Article": "article",
    "German_Plural": "plural",
    "German_Sentence": "german_sentence",
    "English_Word": "english",
    "English_Sentence": "english_sentence",
    "past_participle": "past_participle",
    "simple_past": "simple_past",
    "German_Comparative": "comparative",
    "German_Superlative": "superlative",
}

# Required DB column names that must be non-empty for each table
REQUIRED_FIELDS: dict[str, list[str]] = {
    "verbs": ["type", "english", "german_sentence", "english_sentence"],
    "nouns": ["type", "article", "english", "german_sentence", "english_sentence"],
    "adverbs_adjectives": ["type", "english", "german_sentence", "english_sentence"],
}

# Canonical `type` values each TABLE may contain (validated after lowercasing). Per-table,
# not one global set: "verb" inside nouns.xlsx is just as wrong as a typo like "noum".
# Previously a typo'd Type synced silently to D1 and the app misclassified the word at
# runtime (the ledger/apply paths coerce unknown types) — now it fails the sync like every
# other validation error, with --skip-invalid as the explicit escape hatch.
ALLOWED_TYPES: dict[str, set[str]] = {
    "verbs": {"verb"},
    "nouns": {"noun"},
    "adverbs_adjectives": {"adverb", "adjective"},
}

# A1, A2, B1, B2, C1, C2 with optional .1 or .2 sub-level
VALID_LEVEL = re.compile(r"^(A1|A2|B1|B2|C1|C2)(\.[12])?$")

UPSERT_CHUNK_SIZE = 200
MAX_RETRIES = 3


class ValidationError(Exception):
    """Raised when an Excel file fails structural or row-level validation."""


class _LevelAwareFormatter(logging.Formatter):
    """Plain message for INFO/DEBUG; level-prefixed for WARNING and above."""

    def format(self, record: logging.LogRecord) -> str:
        msg = record.getMessage()
        if record.levelno >= logging.WARNING:
            return f"[{record.levelname}] {msg}"
        return msg


def _setup_logging(verbose: bool, quiet: bool) -> None:
    level = logging.DEBUG if verbose else logging.WARNING if quiet else logging.INFO
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(_LevelAwareFormatter())
    logging.basicConfig(level=level, handlers=[handler], force=True)
    # Keep third-party HTTP libraries quiet even in --verbose mode
    logging.getLogger("httpx").setLevel(logging.WARNING)
    logging.getLogger("httpcore").setLevel(logging.WARNING)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _db_col(header: str) -> str:
    return HEADER_TO_DB.get(header, header.lower())


def _clean(value: Any) -> Any:
    """Strip regular and non-breaking whitespace from strings; return None for empty."""
    if isinstance(value, str):
        return value.replace("\xa0", " ").strip() or None
    return value


_TRUTHY = {"1", "true", "yes", "x", "y"}


def _to_bool01(value: Any) -> int:
    """Coerce a spreadsheet cell to 0/1.

    openpyxl returns numeric cells as floats (a '1' becomes 1.0), so a plain
    string check would miss it. Handle numerics and booleans explicitly, then
    fall back to common string markers; everything else (incl. None, '0', '0.0',
    blank) is 0.
    """
    if value is None:
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return 1 if value != 0 else 0
    return 1 if str(value).strip().lower() in _TRUTHY else 0


_NOUN_ARTICLES = {"der", "die", "das"}


def _capitalize_noun(value: str) -> str:
    """Capitalize a German noun's first letter, preserving an optional leading
    article. 'hund' -> 'Hund', 'die hunde' -> 'die Hunde'. Done once at ingest so
    the app never capitalizes (and mutates) nouns at runtime."""
    text = value.strip()
    if not text:
        return text
    head, sep, tail = text.partition(" ")
    if sep and head.lower() in _NOUN_ARTICLES and tail:
        return f"{head.lower()} {tail[:1].upper()}{tail[1:]}"
    return text[:1].upper() + text[1:]


def compute_id(level: str, word: str) -> str:
    raw = f"{level.lower()}|{word.lower()}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


def compute_content_hash(row: dict[str, Any]) -> str:
    concatenated = "|".join(
        str(row[k]) if row[k] is not None else ""
        for k in sorted(row.keys())
        if k not in ("id", "content_hash")
    )
    return hashlib.sha256(concatenated.encode()).hexdigest()


# ---------------------------------------------------------------------------
# Excel reading and validation
# ---------------------------------------------------------------------------

def read_excel(
    table: str,
    skip_invalid: bool = False,
) -> tuple[list[dict[str, Any]], int, set[str]]:
    """Read and validate the Excel file for a table.

    Returns (valid_rows, skipped_row_count, protected_ids).

    Structural problems (missing file, bad headers, ...) always raise
    ValidationError. Row-level problems raise too by default; with
    skip_invalid=True the offending rows are skipped with a warning instead,
    and the sync continues with the valid rows. protected_ids holds the ids of
    skipped rows whose Level+Word were still computable — the diff must NOT
    delete their existing DB counterparts just because the row was skipped.
    """
    filename, headers = TABLE_CONFIG[table]
    path = DATA_DIR / filename

    if not path.exists():
        raise ValidationError(f"'{filename}' not found at {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        # Multiple sheets → abort: only one sheet is expected
        if len(wb.sheetnames) > 1:
            raise ValidationError(
                f"'{filename}' has multiple sheets: {wb.sheetnames}\n"
                f"  Remove all but one sheet and retry."
            )

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # Read and validate the header row
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValidationError(f"'{filename}' is completely empty.")

        actual = [
            str(c).replace("\xa0", " ").strip() if c is not None else ""
            for c in header_row
        ]
        # Drop trailing empty header cells — openpyxl commonly pads the header row.
        while actual and actual[-1] == "":
            actual.pop()

        # Columns are matched BY NAME, not position (changed 2026-07-12): adding a language
        # is data-only (GL-LANG), and its columns get inserted wherever reads best in the
        # sheet (the Spanish set sits mid-sheet next to its German counterparts). The ingest
        # requires every expected header to EXIST, reads exactly those columns wherever they
        # are, and ignores the rest. Duplicates of an expected header are an error — the
        # ingest could not know which column to trust.
        missing = [h for h in headers if h not in actual]
        if missing:
            raise ValidationError("\n".join([
                f"'{filename}' column mismatch:",
                f"  Expected : {headers}",
                f"  Got      : {actual}",
                f"  Missing  : {missing}",
            ]))
        duplicated = [h for h in headers if actual.count(h) > 1]
        if duplicated:
            raise ValidationError(
                f"'{filename}' has duplicated expected column(s): {duplicated} — "
                f"remove the duplicates so the ingest knows which column to read."
            )
        extra_cols = [c for c in actual if c and c not in headers]
        if extra_cols:
            logger.warning(
                "'%s' has %d unrecognised column(s) beyond the expected set — ignoring: %s",
                filename, len(extra_cols), extra_cols,
            )

        # Sheet position of every expected column (validated present + unique above).
        col_of = {h: actual.index(h) for h in headers}
        level_idx = col_of["Level"]
        word_idx = col_of["German_Word"]
        image_idx = col_of.get("Image", -1)
        free_idx = col_of.get("Free", -1)

        rows: list[dict[str, Any]] = []
        seen_ids: set[str] = set()
        validation_errors: list[str] = []
        skipped_rows = 0
        protected_ids: set[str] = set()

        for row_num, raw_row in enumerate(rows_iter, start=2):  # row 1 is the header
            # Silently skip completely blank rows
            if all(_clean(c) is None for c in raw_row):
                continue

            # Build record — coerce unexpected types (datetime, float) to string.
            # Cells are read through `col_of` (the by-name map), never by list position.
            image_raw: Any = raw_row[image_idx] if 0 <= image_idx < len(raw_row) else None
            record: dict[str, Any] = {}
            for header in headers:
                j = col_of[header]
                db_col = _db_col(header)
                val = _clean(raw_row[j] if j < len(raw_row) else None)
                if val is not None and not isinstance(val, str):
                    val = str(val)
                record[db_col] = val

            # Image is a boolean — handle separately so it stays 0/1
            if table == "nouns":
                record["image"] = _to_bool01(image_raw)

            # Free is a boolean flag gating the paid tier. openpyxl yields 1.0 for
            # a '1' cell, so coerce numerics too (a literal "0" must stay 0).
            free_raw: Any = raw_row[free_idx] if 0 <= free_idx < len(raw_row) else None
            record["free"] = _to_bool01(free_raw)

            # Collect all validation errors for this row before skipping
            row_errors: list[str] = []

            level_raw = _clean(raw_row[level_idx] if level_idx < len(raw_row) else None)
            word_raw = _clean(raw_row[word_idx] if word_idx < len(raw_row) else None)
            level_str = str(level_raw) if level_raw is not None else ""

            if not level_str:
                row_errors.append(f"  Row {row_num}: 'Level' is empty")
            elif not VALID_LEVEL.match(level_str):
                row_errors.append(
                    f"  Row {row_num}: invalid Level {level_str!r} "
                    f"(must be A1/A2/B1/B2/C1/C2, optionally followed by .1 or .2)"
                )

            if not word_raw:
                row_errors.append(f"  Row {row_num}: 'Word' is empty")

            for field in REQUIRED_FIELDS[table]:
                if not record.get(field):
                    label = repr(word_raw) if word_raw else "(unknown word)"
                    row_errors.append(f"  Row {row_num}: required field '{field}' is empty (word={label})")

            # Type must be one of this table's canonical values (checked on the lowercased
            # form, reported with the raw cell so the operator can find it in the sheet).
            # An empty Type is already reported by the required-fields check above.
            type_raw = record.get("type")
            if type_raw and str(type_raw).strip().lower() not in ALLOWED_TYPES[table]:
                label = repr(word_raw) if word_raw else "(unknown word)"
                allowed = "/".join(sorted(ALLOWED_TYPES[table]))
                row_errors.append(
                    f"  Row {row_num}: invalid Type {str(type_raw)!r} "
                    f"— '{table}' allows only {allowed} (word={label})"
                )

            if row_errors:
                validation_errors.extend(row_errors)
                skipped_rows += 1
                # If the row's identity (Level+Word) is intact, remember its id so
                # the diff won't delete the previously-synced DB row for this word.
                if word_raw and VALID_LEVEL.match(level_str):
                    protected_ids.add(compute_id(level_str, str(word_raw)))
                continue

            row_id = compute_id(level_str, str(word_raw))
            if row_id in seen_ids:
                logger.warning(
                    "Row %d: duplicate '%s + %s' after lowercasing — keeping first occurrence",
                    row_num, level_raw, word_raw,
                )
                continue

            seen_ids.add(row_id)
            record["id"] = row_id

            # Normalize once here so the app never does it at runtime:
            #  - canonical lowercase `type` (and `article` for nouns)
            #  - capitalized German nouns + plurals
            # (content_hash is computed AFTER this, so normalized values sync.)
            if record.get("type"):
                record["type"] = str(record["type"]).strip().lower()
            if table == "nouns":
                if record.get("article"):
                    record["article"] = str(record["article"]).strip().lower()
                if record.get("word"):
                    record["word"] = _capitalize_noun(str(record["word"]))
                if record.get("plural"):
                    record["plural"] = _capitalize_noun(str(record["plural"]))

            record["content_hash"] = compute_content_hash(record)
            rows.append(record)

        if validation_errors:
            if not skip_invalid:
                raise ValidationError(
                    f"'{filename}' has {len(validation_errors)} validation error(s):\n"
                    + "\n".join(validation_errors)
                    + "\n  (re-run with --skip-invalid to skip these rows and sync the rest)"
                )
            logger.warning(
                "'%s': skipping %d invalid row(s) (--skip-invalid):\n%s",
                filename, skipped_rows, "\n".join(validation_errors),
            )

        if not rows:
            raise ValidationError(f"'{filename}' contains no valid data rows.")

        return rows, skipped_rows, protected_ids
    finally:
        wb.close()


def find_cross_table_id_collisions(rows_by_table: dict[str, list[dict[str, Any]]]) -> list[str]:
    """Same Level+Word appearing in TWO tables. compute_id has no table component, so such a
    pair shares one id — harmless inside D1 (separate tables) but corrupting everywhere ids are
    global: the audio cache, pack members, and the image decisions store. Zero occurrences in
    the data today; this guard keeps it that way. Returns human-readable problem lines."""
    seen: dict[str, tuple[str, str, str]] = {}
    problems: list[str] = []
    for table, rows in rows_by_table.items():
        for row in rows:
            prev = seen.get(row["id"])
            if prev is not None and prev[0] != table:
                problems.append(
                    f"id {row['id']}: '{prev[2]}' ({prev[1]}, {prev[0]}) collides with "
                    f"'{row['word']}' ({row['level']}, {table})"
                )
            else:
                seen[row["id"]] = (table, row["level"], row["word"])
    return problems


# ---------------------------------------------------------------------------
# HTTP layer with HMAC signing and retry
# ---------------------------------------------------------------------------

def _sign_request(request: httpx.Request) -> None:
    """Sign every outgoing request with HMAC-SHA256.

    Canonical string: METHOD\\nPATH\\nTIMESTAMP\\nSHA256(body_bytes)
    PATH is the request path + query string only (e.g. "/state/verbs") — not the
    full URL. The scheme/host/port are excluded so that proxy normalisation,
    trailing-slash, or default-port differences between client and Worker can
    never silently break signature verification.
    The raw API_KEY never travels over the wire — only the signed digest does.
    The Worker rejects signatures with a timestamp older than 5 minutes.
    """
    timestamp = str(int(time.time()))
    body_hash = hashlib.sha256(request.content).hexdigest()
    path = request.url.raw_path.decode("ascii")  # path + query, e.g. "/sync/verbs"
    canonical = f"{request.method}\n{path}\n{timestamp}\n{body_hash}"
    signature = hmac.new(API_KEY.encode(), canonical.encode(), hashlib.sha256).hexdigest()
    request.headers["X-Timestamp"] = timestamp
    request.headers["X-Signature"] = signature


def _request_with_retry(call) -> httpx.Response:
    """Call call() up to MAX_RETRIES times with exponential backoff.

    Retries on network errors and 5xx responses. Does not retry 4xx errors.
    """
    for attempt in range(MAX_RETRIES):
        try:
            response = call()
            if response.status_code >= 500 and attempt < MAX_RETRIES - 1:
                wait = 2 ** attempt
                logger.warning(
                    "Server error %s, retrying in %ds... (%d/%d)",
                    response.status_code, wait, attempt + 1, MAX_RETRIES,
                )
                time.sleep(wait)
                continue
            return response
        except (httpx.NetworkError, httpx.TimeoutException) as exc:
            if attempt == MAX_RETRIES - 1:
                raise
            wait = 2 ** attempt
            logger.warning(
                "%s, retrying in %ds... (%d/%d)",
                exc.__class__.__name__, wait, attempt + 1, MAX_RETRIES,
            )
            time.sleep(wait)
    raise RuntimeError("retry loop exhausted")  # unreachable


def get_db_state(client: httpx.Client, table: str) -> dict[str, str]:
    response = _request_with_retry(lambda: client.get(f"{WORKER_URL}/state/{table}"))
    response.raise_for_status()
    return response.json()


def post_sync(
    client: httpx.Client,
    table: str,
    upsert: list[dict[str, Any]],
    delete: list[str],
) -> None:
    # Chunk upserts to stay within D1 batch limits; deletes go on the first chunk only
    chunks = [upsert[i: i + UPSERT_CHUNK_SIZE] for i in range(0, max(len(upsert), 1), UPSERT_CHUNK_SIZE)]
    for chunk_idx, chunk in enumerate(chunks):
        chunk_delete = delete if chunk_idx == 0 else []
        response = _request_with_retry(lambda: client.post(
            f"{WORKER_URL}/sync/{table}",
            json={"upsert": chunk, "delete": chunk_delete},
            timeout=60.0,
        ))
        response.raise_for_status()


# ---------------------------------------------------------------------------
# Diff and sync orchestration
# ---------------------------------------------------------------------------

def compute_diff(
    excel_rows: list[dict[str, Any]],
    db_state: dict[str, str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    """Returns (to_insert, to_update, to_delete)."""
    excel_by_id = {row["id"]: row for row in excel_rows}

    to_insert: list[dict[str, Any]] = []
    to_update: list[dict[str, Any]] = []
    for row_id, row in excel_by_id.items():
        if row_id not in db_state:
            to_insert.append(row)
        elif db_state[row_id] != row["content_hash"]:
            to_update.append(row)

    to_delete = [row_id for row_id in db_state if row_id not in excel_by_id]
    return to_insert, to_update, to_delete


# A sync touching more than this many rows (add + update + delete) prompts for
# explicit confirmation before uploading — a guard against an accidental mass change
# (e.g. a malformed Excel that would wipe and re-add everything).
CONFIRM_THRESHOLD = 100


def _confirm_large_change(table: str, total: int, counts: dict[str, int]) -> bool:
    """Ask the operator to confirm a large change. Proceeds only if they type
    'go ahead'; anything else (or a non-interactive stdin) aborts."""
    # Use print/input (not logging) so the prompt shows even under --quiet.
    print(
        f"\n[CONFIRM] '{table}': {total} changes to apply "
        f"({counts['inserted']} add, {counts['updated']} update, {counts['deleted']} delete) "
        f"— exceeds the {CONFIRM_THRESHOLD}-change safety threshold."
    )
    try:
        answer = input("Type 'go ahead' to proceed (anything else aborts): ").strip().lower()
    except (EOFError, KeyboardInterrupt):
        print()
        return False
    return answer == "go ahead"


def sync_table(
    client: httpx.Client,
    table: str,
    excel_rows: list[dict[str, Any]],
    skipped: int,
    protected_ids: set[str],
    dry_run: bool,
) -> dict[str, int]:
    """Sync one table from its pre-read Excel rows. Returns change counts."""
    logger.info("\n[%s]", table)
    logger.info("  %d valid rows in Excel.", len(excel_rows))

    logger.debug("  Fetching DB state...")
    db_state = get_db_state(client, table)
    logger.info("  %d rows in DB.", len(db_state))

    to_insert, to_update, to_delete = compute_diff(excel_rows, db_state)

    # A skipped row must not read as "removed from Excel": keep its DB row as-is.
    # The "*" sentinel (partial --table publish) preserves ALL would-be deletions.
    if "*" in protected_ids:
        preserved = list(to_delete)
    else:
        preserved = [row_id for row_id in to_delete if row_id in protected_ids]
    if preserved:
        to_delete = [row_id for row_id in to_delete if row_id not in protected_ids]
        logger.warning(
            "  %d DB row(s) correspond to skipped Excel rows — preserving them, not deleting.",
            len(preserved),
        )

    counts = {
        "inserted": len(to_insert),
        "updated": len(to_update),
        "deleted": len(to_delete),
        "skipped": skipped,
    }

    logger.info(
        "  Diff: %d to add, %d to update, %d to delete.",
        counts["inserted"], counts["updated"], counts["deleted"],
    )

    if not to_insert and not to_update and not to_delete:
        logger.info("  Already in sync.")
        return counts

    if dry_run:
        logger.info("  [DRY RUN] No changes written.")
        return counts

    total = counts["inserted"] + counts["updated"] + counts["deleted"]
    if total > CONFIRM_THRESHOLD and not _confirm_large_change(table, total, counts):
        logger.warning("  Aborted by user — '%s' NOT uploaded.", table)
        counts["aborted"] = True
        return counts

    post_sync(client, table, to_insert + to_update, to_delete)
    logger.info("  Done.")
    return counts


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

V2_PUBLISH_ORDER = ("verbs", "nouns", "adverbs_adjectives", "translations", "id_aliases")


def _load_environment(name: str) -> "envs.Environment":
    """Resolve the target environment (sync/envs.py) and point the HTTP layer at it."""
    global WORKER_URL, API_KEY
    env = envs.load_environment(name)
    WORKER_URL = env.worker_url
    API_KEY = env.api_key
    return env


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Publish the vocabulary spreadsheets to an environment's content database.",
    )
    parser.add_argument(
        "--env",
        choices=envs.environment_names(),
        default=envs.DEFAULT_ENV,
        help="Target environment (default: %(default)s). prod additionally requires "
             "the typed confirmation (MS2-FR-30).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would change without writing anything to the DB.",
    )
    parser.add_argument(
        "--skip-invalid",
        action="store_true",
        help="Skip rows that fail row-level validation (empty required fields, "
             "bad Level, ...) and sync the remaining valid rows, instead of "
             "failing the whole table. Skipped words keep their existing DB row "
             "if they have one. Structural errors still fail the table.",
    )
    parser.add_argument(
        "--table",
        choices=list(V2_TABLES.keys()),
        metavar="TABLE",
        help=f"Publish only one word table (its translations and aliases still "
             f"publish). Choices: {', '.join(V2_TABLES.keys())}.",
    )
    verbosity = parser.add_mutually_exclusive_group()
    verbosity.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Show debug-level detail (per-step progress).",
    )
    verbosity.add_argument(
        "-q", "--quiet",
        action="store_true",
        help="Only show warnings and errors (the final summary still prints).",
    )
    args = parser.parse_args()

    _setup_logging(args.verbose, args.quiet)

    try:
        env = _load_environment(args.env)
    except envs.EnvironmentError_ as exc:
        logger.error("%s", exc)
        sys.exit(1)
    logger.info("→ environment: %s (%s)", env.name, env.worker_url)

    word_tables = [args.table] if args.table else list(V2_TABLES.keys())

    if args.dry_run:
        logger.info("[DRY RUN] No changes will be written to the DB.")

    totals: dict[str, int] = {"inserted": 0, "updated": 0, "deleted": 0, "skipped": 0}
    failures: list[str] = []
    aborted: list[str] = []

    # Read + validate ALL word tables first: the aggregated dataset feeds the
    # global-uniqueness guards and the coverage report before anything is written.
    datasets: dict[str, "dataset_mod.TableDataset"] = {}
    for table in word_tables:
        try:
            datasets[table] = dataset_mod.read_dataset(table, skip_invalid=args.skip_invalid)
        except dataset_mod.ValidationError as exc:
            logger.error("%s: %s", table, exc)
            failures.append(table)

    if not datasets:
        logger.error("Nothing readable — aborting.")
        sys.exit(1)

    # Global guards across the whole dataset: a v2 id must be unique across ALL
    # word tables (ids are global in media/packs/approvals), and a legacy id may
    # alias only one v2 id.
    id_owner: dict[str, str] = {}
    alias_owner: dict[str, str] = {}
    guard_problems: list[str] = []
    for table, ds in datasets.items():
        for row in ds.core:
            prev = id_owner.get(row["id"])
            if prev is not None:
                guard_problems.append(f"id {row['id']} ({row['word']!r}) in both {prev} and {table}")
            id_owner[row["id"]] = table
        for alias in ds.aliases:
            prev = alias_owner.get(alias["id"])
            if prev is not None:
                guard_problems.append(
                    f"legacy id {alias['id']} maps to two v2 ids ({prev} / {alias['new_id']})"
                )
            alias_owner[alias["id"]] = alias["new_id"]
    if guard_problems:
        logger.error(
            "Global identity guard failed — nothing was synced:\n  %s",
            "\n  ".join(guard_problems),
        )
        sys.exit(1)

    # Coverage preview (LG-FR-14) — always shown before anything uploads.
    print("\nLanguage coverage:")
    for line in dataset_mod.coverage_report(datasets):
        print(line)
    print()

    # Production gate (MS2-FR-30): typed confirmation, refused non-interactively.
    if not args.dry_run:
        try:
            envs.confirm_production(env, action="publish vocabulary")
        except envs.EnvironmentError_ as exc:
            logger.error("%s", exc)
            sys.exit(1)

    # Assemble the five physical tables. Skipped rows protect their previously
    # published counterparts everywhere: core (v2 id), translations (v2 id per
    # language), and aliases (legacy id).
    all_langs = list(dataset_mod.LANGUAGES)
    plan: list[tuple[str, list[dict[str, Any]], int, set[str]]] = []
    for table in V2_PUBLISH_ORDER:
        if table in datasets:
            ds = datasets[table]
            plan.append((table, ds.core, ds.skipped, set(ds.protected)))
    if any(t in datasets for t in V2_TABLES):
        translations = [r for ds in datasets.values() for r in ds.translations]
        aliases = [r for ds in datasets.values() for r in ds.aliases]
        protected_translations = {
            f"{pid}:{code}"
            for ds in datasets.values() for pid in ds.protected for code in all_langs
        }
        protected_aliases = {a for ds in datasets.values() for a in ds.protected_aliases}
        if args.table:
            logger.warning(
                "--table publishes only '%s' word rows, but translations/id_aliases "
                "sync from the SAME partial read — other tables' translations are "
                "protected from deletion only via their skipped/protected sets. "
                "Prefer full publishes.", args.table,
            )
            # Partial read: the diff cannot distinguish "row absent because its
            # table wasn't read" from "row removed" — so delete nothing.
            protected_translations |= {"*"}  # sentinel handled in sync_table
            protected_aliases |= {"*"}
        plan.append(("translations", translations, 0, protected_translations))
        plan.append(("id_aliases", aliases, 0, protected_aliases))

    with httpx.Client(event_hooks={"request": [_sign_request]}) as client:
        for table, rows, skipped, protected_ids in plan:
            try:
                result = sync_table(client, table, rows, skipped, protected_ids, dry_run=args.dry_run)
            except httpx.HTTPError as exc:
                logger.error("%s: request failed: %s", table, exc)
                failures.append(table)
                continue
            if result.get("aborted"):
                aborted.append(table)
                continue
            for key in totals:
                totals[key] += result[key]

    dry_label = " (dry run — no changes written)" if args.dry_run else ""
    print(f"\n{'─' * 44}")
    print(f"Summary{dry_label}")
    print(f"  Added   : {totals['inserted']}")
    print(f"  Updated : {totals['updated']}")
    print(f"  Deleted : {totals['deleted']}")
    if totals["skipped"]:
        print(f"  Skipped : {totals['skipped']} invalid row(s) — NOT synced (see warnings above)")
    if aborted:
        print(f"  Aborted : {len(aborted)} table(s) — {', '.join(aborted)} (not uploaded)")
    if failures:
        print(f"  Failed  : {len(failures)} table(s) — {', '.join(failures)}")
    print(f"{'─' * 44}")

    if failures or aborted:
        sys.exit(1)
    if not args.dry_run:
        print("Sync complete.")


if __name__ == "__main__":
    main()
