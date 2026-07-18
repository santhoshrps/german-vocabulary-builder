"""
Tests for the targeted media-replacement mechanism.

The single most important invariant is HASH STABILITY: a word with no override must hash
byte-identically to a world where the replacement feature does not exist — otherwise every
run would re-ship the entire audio set to every installed app.

Run:  cd sync && .venv/bin/python tests/test_media_replacement.py
"""

from __future__ import annotations

import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

import audio_engine
import audio_overrides
import audio_sync
import image_decisions
import media_replace
import sync


def _noun(level, word, article="der", plural="", sentence="Der Satz.", free=0):
    return {
        "id": sync.compute_id(level, word), "level": level, "word": word, "type": "noun",
        "article": article, "plural": plural, "german_sentence": sentence,
        "english": "x", "english_sentence": "x", "free": free, "image": 0, "capital": None,
    }


def _verb(level, word, sentence="Ein Satz."):
    return {
        "id": sync.compute_id(level, word), "level": level, "word": word, "type": "verb",
        "german_sentence": sentence, "english": "x", "english_sentence": "x", "free": 0,
        "capital": None, "ich": None, "du": None, "er_sie_es": None, "wir": None, "ihr": None,
        "sie_sie": None, "past_participle": None, "simple_past": None,
    }


VOCAB = {
    "nouns": [
        _noun("A1.1", "Hund", "der", "Hunde"),
        _noun("A1.1", "Nachricht", "die"),
        _noun("A2.2", "Nachricht", "die"),
    ],
    "verbs": [_verb("A1.1", "laufen")],
    "adverbs_adjectives": [{
        "id": sync.compute_id("A1.1", "schnell"), "level": "A1.1", "word": "schnell",
        "type": "adjective", "german_sentence": "Er ist schnell.", "english": "x",
        "english_sentence": "x", "free": 0, "capital": None, "comparative": None, "superlative": None,
    }],
}
HUND = VOCAB["nouns"][0]


class TestHashStability(unittest.TestCase):
    def test_take_zero_is_the_legacy_hash(self):
        self.assertEqual(audio_engine.audio_hash("der Hund", "de-DE-ConradNeural"),
                         audio_engine.audio_hash("der Hund", "de-DE-ConradNeural", 0))

    def test_take_changes_hash(self):
        base = audio_engine.audio_hash("der Hund", "de-DE-ConradNeural")
        t1 = audio_engine.audio_hash("der Hund", "de-DE-ConradNeural", 1)
        t2 = audio_engine.audio_hash("der Hund", "de-DE-ConradNeural", 2)
        self.assertNotEqual(base, t1)
        self.assertNotEqual(t1, t2)

    def test_descriptors_without_overrides_are_untouched(self):
        baseline = audio_sync._row_descriptors("nouns", HUND)
        again = audio_sync._row_descriptors("nouns", HUND)
        audio_sync.apply_overrides("nouns", HUND, again, {})
        self.assertEqual(baseline, again)

    def test_override_touches_only_its_clip(self):
        descs = audio_sync._row_descriptors("nouns", HUND)
        baseline = {d["id"]: d["audio_hash"] for d in descs}
        overrides = {HUND["id"]: {"take": 1}}   # singular only
        audio_sync.apply_overrides("nouns", HUND, descs, overrides)
        for d in descs:
            if d["id"] == HUND["id"]:
                self.assertNotEqual(baseline[d["id"]], d["audio_hash"])
            else:
                self.assertEqual(baseline[d["id"]], d["audio_hash"])


class TestVoiceRotation(unittest.TestCase):
    def test_rotation_always_leaves_the_base_voice(self):
        pool = audio_engine.MASCULINE_VOICES
        enabled = [v for v in pool if v not in audio_engine.DISABLED_VOICES]
        base = enabled[0]
        for take in range(1, len(enabled)):
            self.assertNotEqual(audio_engine.rotated_voice(base, pool, take), base)

    def test_rotation_is_deterministic_and_take_zero_is_base(self):
        pool = audio_engine.FEMININE_VOICES
        base = [v for v in pool if v not in audio_engine.DISABLED_VOICES][0]
        self.assertEqual(audio_engine.rotated_voice(base, pool, 0), base)
        self.assertEqual(audio_engine.rotated_voice(base, pool, 3),
                         audio_engine.rotated_voice(base, pool, 3))

    def test_rotation_never_picks_a_disabled_voice(self):
        pool = audio_engine.ALL_VOICES
        base = [v for v in pool if v not in audio_engine.DISABLED_VOICES][0]
        for take in range(1, 40):
            self.assertNotIn(audio_engine.rotated_voice(base, pool, take),
                             audio_engine.DISABLED_VOICES)

    def test_pool_matches_gender(self):
        self.assertEqual(audio_engine.voice_pool_for("nouns", HUND, "singular"),
                         audio_engine.MASCULINE_VOICES)
        self.assertEqual(audio_engine.voice_pool_for("nouns", HUND, "sentence"),
                         audio_engine.ALL_VOICES)
        self.assertEqual(audio_engine.voice_pool_for("verbs", VOCAB["verbs"][0], "singular"),
                         audio_engine.ALL_VOICES)


class TestOverrideApply(unittest.TestCase):
    def _desc(self):
        return audio_sync._row_descriptors("nouns", HUND)[0]  # singular: "der Hund"

    def test_hint_respells_the_word_everywhere(self):
        d = audio_sync._row_descriptors("nouns", _noun("A1.1", "Hund", "der", "Hunde",
                                                       sentence="Der Hund bellt laut."))[2]
        audio_overrides.apply(d, {"take": 1, "hint": "Hunt"}, audio_engine.ALL_VOICES, "Hund")
        self.assertIn("Hunt", d["text"])
        self.assertNotIn("Hund ", d["text"])

    def test_voice_pin_beats_rotation(self):
        d = self._desc()
        audio_overrides.apply(d, {"take": 2, "voice": "de-DE-KatjaNeural"},
                              audio_engine.MASCULINE_VOICES, "Hund")
        self.assertEqual(d["voice"], "de-DE-KatjaNeural")
        self.assertEqual(d["audio_hash"], audio_engine.audio_hash(d["text"], "de-DE-KatjaNeural", 2))

    def test_take_rotates_within_pool(self):
        d = self._desc()
        base_voice = d["voice"]
        audio_overrides.apply(d, {"take": 1}, audio_engine.MASCULINE_VOICES, "Hund")
        self.assertNotEqual(d["voice"], base_voice)
        self.assertIn(d["voice"], audio_engine.MASCULINE_VOICES)


class TestCollisionGuard(unittest.TestCase):
    def test_clean_data_passes(self):
        self.assertEqual(sync.find_cross_table_id_collisions(
            {t: rows for t, rows in VOCAB.items()}), [])

    def test_cross_table_collision_is_reported(self):
        bad = {"nouns": [_noun("A1.1", "Essen")], "verbs": [_verb("A1.1", "essen")]}
        problems = sync.find_cross_table_id_collisions(bad)
        self.assertEqual(len(problems), 1)
        self.assertIn("essen", problems[0].lower())


class TestImageReplacementSemantics(unittest.TestCase):
    def test_approved_record_is_kept_and_marked(self):
        store = {"abc": {"status": "approved", "content_hash": "oldhash", "input_fingerprint": "f"}}
        prev = image_decisions.request_replacement(store, "abc")
        self.assertEqual(prev, "oldhash")
        self.assertEqual(store["abc"]["status"], "approved")     # zero-gap: still shipping
        self.assertTrue(store["abc"]["replace_requested"])
        self.assertFalse(image_decisions.is_settled_current(
            store, {"id": "abc", "english": "", "word": "", "german_sentence": ""}))

    def test_non_approved_record_is_forgotten(self):
        store = {"abc": {"status": "review", "input_fingerprint": "f"}}
        self.assertIsNone(image_decisions.request_replacement(store, "abc"))
        self.assertNotIn("abc", store)

    def test_new_approval_clears_the_request(self):
        store = {"abc": {"status": "approved", "content_hash": "old", "replace_requested": True}}
        noun = {"id": "abc", "english": "dog", "word": "Hund", "german_sentence": "s"}
        image_decisions.record_approved(store, noun, source="s", source_id="i", url="", license="l",
                                        kind="photo", content_hash="new", approved_by="human",
                                        today="2026-07-18")
        self.assertNotIn("replace_requested", store["abc"])
        self.assertEqual(store["abc"]["content_hash"], "new")


class TestZeroGapSourcePass(unittest.TestCase):
    """image_sync._source_pass with a replacement requested on an APPROVED noun: the approved
    record (and thus the shipped image) must survive candidate generation, must not regenerate
    while candidates are queued, and must be kept when a round produces nothing."""

    def setUp(self):
        import types
        import image_sync
        self.image_sync = image_sync
        self.noun = dict(HUND)
        self.fingerprint = image_decisions.input_fingerprint(self.noun)
        self.store = {self.noun["id"]: {
            "status": "approved", "content_hash": "oldhash",
            "input_fingerprint": self.fingerprint, "replace_requested": True,
        }}
        self.queue: dict = {}
        cand = types.SimpleNamespace(
            master=b"IMG", content_hash="newcand",
            candidate=types.SimpleNamespace(source="gen", source_id="x", page_url="", image_url="",
                                            license="generated"),
            kind="photo", verifier=None, clip=None)
        self.review_outcome = types.SimpleNamespace(status="review", candidates=[cand], chosen=None)
        self.none_outcome = types.SimpleNamespace(status="none", candidates=[], chosen=None)
        self.calls = 0
        for target, repl in [
            ("_write_master", lambda master, h: None),
            ("_save_review_queue", lambda q: None),
        ]:
            p = mock.patch.object(image_sync, target, repl)
            p.start()
            self.addCleanup(p.stop)
        p = mock.patch.object(image_decisions, "save", lambda store: None)
        p.start()
        self.addCleanup(p.stop)

    def _run_pass(self, outcome):
        def fake_process(noun, **kwargs):
            self.calls += 1
            return outcome
        with mock.patch.object(self.image_sync.image_engine, "process_noun", fake_process):
            self.image_sync._source_pass([self.noun], self.store, self.queue, {},
                                         client=None, bucket=None, dry_run=False, limit=0)

    def test_review_round_keeps_the_approved_record(self):
        self._run_pass(self.review_outcome)
        rec = self.store[self.noun["id"]]
        self.assertEqual(rec["status"], "approved")            # old image still ships
        self.assertEqual(rec["content_hash"], "oldhash")
        self.assertEqual(rec["replace_fingerprint"], self.fingerprint)
        self.assertIn(self.noun["id"], self.queue)             # candidates await review
        self.assertEqual(self.calls, 1)
        # Second run: candidates queued for current content → must NOT regenerate.
        self._run_pass(self.review_outcome)
        self.assertEqual(self.calls, 1)

    def test_failed_round_keeps_the_current_image(self):
        self._run_pass(self.none_outcome)
        rec = self.store[self.noun["id"]]
        self.assertEqual(rec["status"], "approved")
        self.assertEqual(rec["content_hash"], "oldhash")
        self.assertNotIn("replace_requested", rec)             # request settled, image kept
        self.assertNotIn(self.noun["id"], self.queue)


class TestBacklogTool(unittest.TestCase):
    """End-to-end lifecycle of media_replace.py against fixture vocab + a temp backlog sheet."""

    HEADERS = ["Word", "Type", "Level (auto)", "Replace_Audio", "Replace_Image", "Audio_Variants",
               "Voice", "Pronunciation_Hint", "Image_Note", "Status", "Remarks"]

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="media_replace_test_"))
        self.addCleanup(shutil.rmtree, self.tmp, ignore_errors=True)
        self.sheet = self.tmp / "media_replacements.xlsx"
        self.preview = self.tmp / "media_preview"
        self.overrides_file = self.tmp / "audio_overrides.json"
        self.index_file = self.tmp / "audio_index.json"
        self.store: dict = {}
        self.queue: dict = {}
        self.opts: dict = {}

        patches = [
            mock.patch.object(media_replace, "SHEET_PATH", self.sheet),
            mock.patch.object(media_replace, "PREVIEW_DIR", self.preview),
            mock.patch.object(media_replace, "STATE_PATH", self.preview / "state.json"),
            mock.patch.object(audio_sync, "INDEX_PATH", self.index_file),
            mock.patch.object(sync, "read_excel",
                              lambda table, skip_invalid=False: (VOCAB[table], 0, set())),
            mock.patch.object(audio_engine, "synthesize",
                              lambda text, voice, path: Path(path).write_bytes(b"FAKE " + voice.encode())),
            mock.patch.object(audio_overrides, "load", lambda: self._load_json(self.overrides_file)),
            mock.patch.object(audio_overrides, "save",
                              lambda store: self.overrides_file.write_text(json.dumps(store))),
            mock.patch.object(image_decisions, "load", lambda: self.store),
            mock.patch.object(image_decisions, "save", lambda store: None),
            mock.patch.object(image_decisions, "load_review_queue", lambda: self.queue),
            mock.patch.object(image_decisions, "save_review_queue", lambda q: None),
            mock.patch.object(image_decisions, "load_prompt_opts", lambda: self.opts),
            mock.patch.object(image_decisions, "save_prompt_opts", lambda o: None),
        ]
        for p in patches:
            p.start()
            self.addCleanup(p.stop)

    @staticmethod
    def _load_json(path: Path) -> dict:
        return json.loads(path.read_text()) if path.exists() else {}

    def _write_sheet(self, rows: list[dict]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(self.HEADERS)
        for r in rows:
            ws.append([r.get(h, "") for h in self.HEADERS])
        wb.save(self.sheet)

    def _run(self, *argv, expect_exit=None):
        with mock.patch.object(sys, "argv", ["media_replace.py", *argv]):
            if expect_exit is None:
                media_replace.main()
            else:
                with self.assertRaises(SystemExit) as ctx:
                    media_replace.main()
                self.assertEqual(ctx.exception.code, expect_exit)

    def _statuses(self) -> list[str]:
        wb = openpyxl.load_workbook(self.sheet)
        ws = wb.active
        out = [str(ws.cell(r, 10).value or "") for r in range(2, ws.max_row + 1)]
        wb.close()
        return out

    def _levels(self) -> list[str]:
        wb = openpyxl.load_workbook(self.sheet)
        ws = wb.active
        out = [str(ws.cell(r, 3).value or "") for r in range(2, ws.max_row + 1)]
        wb.close()
        return out

    # -- lifecycle ----------------------------------------------------------

    def test_full_audio_lifecycle(self):
        self._write_sheet([{"Word": "Hund", "Type": "noun", "Replace_Audio": "x"}])
        self._run()
        status = self._statuses()[0]
        self.assertIn("PREVIEW take 1", status)
        self.assertEqual(self._levels()[0], "A1.1")           # Level auto-filled
        clips = list(self.preview.glob("Hund__*.m4a"))
        self.assertEqual(len(clips), 3)                       # singular + plural + sentence

        self._run("--approve")
        self.assertIn("approved take 1", self._statuses()[0])
        committed = self._load_json(self.overrides_file)
        self.assertEqual(len(committed), 3)
        self.assertEqual(list(self.preview.glob("Hund__*.m4a")), [])   # previews cleaned up

        # Simulate audio_sync publishing: its collect_words applies the same overrides.
        words = audio_sync.collect_words(committed)
        self.index_file.write_text(json.dumps({w["id"]: w["audio_hash"] for w in words}))
        self._run()
        self.assertIn("published ✓", self._statuses()[0])

    def test_rejecting_a_preview_advances_take_and_voice(self):
        self._write_sheet([{"Word": "Hund", "Type": "noun", "Replace_Audio": "x",
                            "Audio_Variants": "singular"}])
        self._run()
        first = next(self.preview.glob("Hund__singular_*.m4a"))
        # Reject: clear Status, run again → take 2, a different file name (voice/take differ).
        wb = openpyxl.load_workbook(self.sheet)
        wb.active.cell(2, 10).value = None
        wb.save(self.sheet)
        self._run()
        second = next(self.preview.glob("Hund__singular_*.m4a"))
        self.assertIn("take2", second.name)
        self.assertNotEqual(first.name, second.name)

    def test_full_image_lifecycle_zero_gap(self):
        hund_id = HUND["id"]
        fingerprint = image_decisions.input_fingerprint(HUND)
        self.store[hund_id] = {"status": "approved", "content_hash": "oldhash",
                               "input_fingerprint": fingerprint}
        self._write_sheet([{"Word": "die Hund", "Type": "noun", "Replace_Image": "x",
                            "Image_Note": "show the whole dog"}])
        self._run()
        self.assertIn("image: queued", self._statuses()[0])
        self.assertTrue(self.store[hund_id]["replace_requested"])      # old image still approved
        self.assertEqual(self.store[hund_id]["status"], "approved")
        self.assertEqual(self.opts[hund_id]["note"], "show the whole dog")

        # image_sync generated candidates → queued for review.
        self.queue[hund_id] = {"word": "Hund", "candidates": []}
        self.store[hund_id]["replace_fingerprint"] = fingerprint
        self._run()
        self.assertIn("awaiting review", self._statuses()[0])

        # Reviewer picked a new image.
        image_decisions.record_approved(self.store, HUND, source="s", source_id="i", url="",
                                        license="l", kind="photo", content_hash="newhash",
                                        approved_by="human", today="2026-07-18")
        self.queue.pop(hund_id)
        self._run()
        self.assertIn("image: done ✓", self._statuses()[0])

    # -- validation ---------------------------------------------------------

    def test_validation_errors(self):
        self._write_sheet([
            {"Word": "Nachricht", "Type": "noun", "Replace_Audio": "x"},          # ambiguous
            {"Word": "laufen", "Type": "verb", "Replace_Image": "x"},             # image on verb
            {"Word": "laufen", "Type": "verb", "Replace_Audio": "x",
             "Audio_Variants": "plural"},                                          # verbs: no plural
            {"Word": "Katze", "Type": "noun", "Replace_Audio": "x"},              # not in vocab
            {"Word": "schnell", "Type": "adverb", "Replace_Audio": "x"},          # wrong type
            {"Word": "Hund", "Type": "noun"},                                      # nothing marked
        ])
        self._run(expect_exit=1)
        statuses = self._statuses()
        self.assertIn("ambiguous", statuses[0])
        self.assertIn("A1.1, A2.2", statuses[0])
        self.assertIn("not enabled for verbs", statuses[1])
        self.assertIn("no plural clip", statuses[2])
        self.assertIn("not found", statuses[3])
        self.assertIn("not found", statuses[4])
        self.assertIn("mark Replace_Audio", statuses[5])

    def test_level_pin_resolves_ambiguity(self):
        self._write_sheet([{"Word": "Nachricht", "Type": "noun", "Level (auto)": "A2.2",
                            "Replace_Audio": "x", "Audio_Variants": "singular"}])
        self._run()
        self.assertIn("PREVIEW", self._statuses()[0])

    def test_duplicate_rows_rejected(self):
        self._write_sheet([
            {"Word": "Hund", "Type": "noun", "Replace_Audio": "x", "Audio_Variants": "singular"},
            {"Word": "der Hund", "Type": "noun", "Replace_Audio": "x"},
        ])
        self._run(expect_exit=1)
        self.assertIn("duplicate", self._statuses()[1])

    def test_dry_run_changes_nothing(self):
        self._write_sheet([{"Word": "Hund", "Type": "noun", "Replace_Audio": "x",
                            "Replace_Image": "x"}])
        self._run("--dry-run")
        self.assertEqual(self._statuses()[0], "")              # sheet untouched
        self.assertEqual(list(self.preview.glob("*.m4a")), []) # nothing synthesized
        self.assertEqual(self.store, {})                       # no image mutation
        self.assertFalse(self.overrides_file.exists())


if __name__ == "__main__":
    unittest.main(verbosity=2)
